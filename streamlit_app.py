
import os
import tempfile
import hashlib
import base64
from datetime import datetime, date, time
from zoneinfo import ZoneInfo
from pathlib import Path

import pandas as pd
import streamlit as st
from openpyxl import load_workbook

from planner_core import Database


APP_TITLE = "Electro-Dip Production Planner with WIP"
BASE_DIR = Path(__file__).resolve().parent
DB_PATH = BASE_DIR / "electro_dip_persistent.db"
TEMPLATE_PATH = BASE_DIR / "Electro_Dip_WIP_Import_Template_V7.xlsx"

st.set_page_config(page_title=APP_TITLE, page_icon="🏭", layout="wide")


def get_db():
    return Database(DB_PATH)



db = get_db()


def make_operator_slip_pdf_bytes(selected_date=None, selected_machine=None):
    with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
        path = Path(tmp.name)
    try:
        db.create_operator_slips_pdf(
            path,
            due_date=selected_date,
            machine_name=selected_machine,
        )
        return path.read_bytes()
    finally:
        path.unlink(missing_ok=True)


def make_production_plan_excel_bytes():
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        path = Path(tmp.name)
    try:
        db.export_production_plan_excel(path)
        return path.read_bytes()
    finally:
        path.unlink(missing_ok=True)


REQUIRED_SHEETS = {
    "Customer_Schedules": [
        "Schedule ID", "Customer Name", "Part Name",
        "Customer Required Qty", "Required Delivery Date",
        "Required Delivery Time", "Priority",
    ],
    "Stock_Demand": [
        "Part Name", "Current FG Stock", "Minimum Stock", "Remarks",
    ],
    "Batch_Config": [
        "Part Name", "Production Batch Qty", "Transportation Batch Qty",
    ],
    "Process_BOM": [
        "Part Name", "Process Sequence", "Operation Name", "Process Type",
        "Cycle Time Sec/Part", "Setup Time Min",
        "Outsource Lead Time Hours", "Qty Multiplier",
        "Scrap Allowance %",
    ],
    "Machine_Recommendations": [
        "Part Name", "Operation Name",
    ],
}

OPTIONAL_SHEETS = {
    "Machine_Downtime": [
        "Machine Name", "Unavailable Start Date", "Unavailable Start Time",
        "Unavailable End Date", "Unavailable End Time", "Reason",
    ],
    "Shifts": ["Shift Name", "Start Time", "End Time", "Active (Y/N)"],
    "Breaks": [
        "Shift Name", "Break Name", "Break Start Time", "Break End Time",
    ],
    "Holidays": ["Holiday Date", "Holiday Name"],
    "Weekly_Offs": ["Day Number", "Day Name", "Is Weekly Off (Y/N)"],
    "Opening_WIP": [
        "Entry Date", "Part Name", "Process Sequence", "Operation Name",
        "Actual Qty", "Rejected Qty", "Good Qty", "Machine",
        "Operator", "Remarks",
    ],
}


def _diag_value_blank(value):
    return value is None or str(value).strip() == ""


def _diag_number(value):
    if _diag_value_blank(value):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _diag_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        try:
            return (datetime(1899, 12, 30) + pd.to_timedelta(float(value), unit="D")).date()
        except Exception:
            return None
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d-%b-%Y", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    return None


def _diag_time(value):
    if isinstance(value, datetime):
        return value.time()
    if isinstance(value, time):
        return value
    if isinstance(value, (int, float)):
        return True
    text = str(value).strip()
    for fmt in ("%H:%M", "%H:%M:%S", "%I:%M %p"):
        try:
            datetime.strptime(text, fmt)
            return True
        except ValueError:
            pass
    return None


def _add_diagnostic(items, severity, sheet, row, column, reason, value=""):
    items.append({
        "Severity": severity,
        "Sheet": sheet,
        "Row": row,
        "Column": column,
        "Reason": reason,
        "Value": "" if value is None else str(value),
    })


def validate_import_template(file_source):
    """
    Validate an Electro-Dip import workbook before importing.

    Returns:
      valid_for_import: True when no ERROR diagnostics exist
      diagnostics: sheet/row/column-level findings
      sheet_summary: validation status and populated-row counts
    """
    if hasattr(file_source, "seek"):
        file_source.seek(0)

    workbook = load_workbook(
        file_source,
        data_only=True,
        read_only=True,
    )

    diagnostics = []
    sheet_summary = []
    workbook_sheets = set(workbook.sheetnames)

    # Sheet and heading validation.
    for sheet_name, required_headers in REQUIRED_SHEETS.items():
        if sheet_name not in workbook_sheets:
            _add_diagnostic(
                diagnostics, "ERROR", sheet_name, 0, "",
                "Required worksheet is missing.",
            )
            sheet_summary.append({
                "Sheet": sheet_name,
                "Status": "Missing",
                "Populated Rows": 0,
            })
            continue

        ws = workbook[sheet_name]
        headers = [
            str(cell.value).strip() if cell.value is not None else ""
            for cell in ws[2]
        ]
        missing_headers = [
            header for header in required_headers if header not in headers
        ]
        for header in missing_headers:
            _add_diagnostic(
                diagnostics, "ERROR", sheet_name, 2, header,
                "Required column heading is missing.",
            )

    for sheet_name, expected_headers in OPTIONAL_SHEETS.items():
        if sheet_name not in workbook_sheets:
            _add_diagnostic(
                diagnostics, "WARNING", sheet_name, 0, "",
                "Optional worksheet is missing.",
            )
            continue
        ws = workbook[sheet_name]
        headers = [
            str(cell.value).strip() if cell.value is not None else ""
            for cell in ws[2]
        ]
        for header in expected_headers:
            if header not in headers:
                _add_diagnostic(
                    diagnostics, "WARNING", sheet_name, 2, header,
                    "Expected optional column heading is missing.",
                )

    # Stop row scanning only after sheet structure checks.
    populated_counts = {}

    def rows_until_blank(ws, max_blank_run=100):
        blank_run = 0
        for row_number, values in enumerate(
            ws.iter_rows(min_row=3, values_only=True), start=3
        ):
            if all(_diag_value_blank(value) for value in values):
                blank_run += 1
                if blank_run >= max_blank_run:
                    break
                continue
            blank_run = 0
            yield row_number, values

    schedule_parts = set()
    process_parts = set()
    stock_parts = set()
    batch_parts = set()
    machine_routes = set()

    # Customer schedules.
    if "Customer_Schedules" in workbook_sheets:
        ws = workbook["Customer_Schedules"]
        seen_schedule_ids = {}
        count = 0
        for row_no, row in rows_until_blank(ws):
            count += 1
            values = list(row) + [None] * 7
            schedule_id, customer, part, qty, due_date, due_time, priority = values[:7]

            mandatory = {
                "Schedule ID": schedule_id,
                "Customer Name": customer,
                "Part Name": part,
                "Customer Required Qty": qty,
                "Required Delivery Date": due_date,
                "Required Delivery Time": due_time,
            }
            for column, value in mandatory.items():
                if _diag_value_blank(value):
                    _add_diagnostic(
                        diagnostics, "ERROR", "Customer_Schedules",
                        row_no, column, "Mandatory value is blank.", value,
                    )

            sid = str(schedule_id).strip() if not _diag_value_blank(schedule_id) else ""
            if sid:
                if sid in seen_schedule_ids:
                    _add_diagnostic(
                        diagnostics, "ERROR", "Customer_Schedules",
                        row_no, "Schedule ID",
                        f"Duplicate Schedule ID; first found at row {seen_schedule_ids[sid]}.",
                        sid,
                    )
                else:
                    seen_schedule_ids[sid] = row_no

            if not _diag_value_blank(part):
                schedule_parts.add(str(part).strip())

            qty_num = _diag_number(qty)
            if not _diag_value_blank(qty) and qty_num is None:
                _add_diagnostic(
                    diagnostics, "ERROR", "Customer_Schedules",
                    row_no, "Customer Required Qty",
                    "Quantity must be numeric.", qty,
                )
            elif qty_num is not None and qty_num <= 0:
                _add_diagnostic(
                    diagnostics, "ERROR", "Customer_Schedules",
                    row_no, "Customer Required Qty",
                    "Quantity must be greater than zero.", qty,
                )

            if not _diag_value_blank(due_date) and _diag_date(due_date) is None:
                _add_diagnostic(
                    diagnostics, "ERROR", "Customer_Schedules",
                    row_no, "Required Delivery Date",
                    "Invalid date value.", due_date,
                )

            if not _diag_value_blank(due_time) and _diag_time(due_time) is None:
                _add_diagnostic(
                    diagnostics, "ERROR", "Customer_Schedules",
                    row_no, "Required Delivery Time",
                    "Invalid time value.", due_time,
                )

            priority_num = _diag_number(priority)
            if not _diag_value_blank(priority) and priority_num is None:
                _add_diagnostic(
                    diagnostics, "WARNING", "Customer_Schedules",
                    row_no, "Priority",
                    "Priority should be numeric; default priority will be used.",
                    priority,
                )

        populated_counts["Customer_Schedules"] = count

    # Stock master.
    if "Stock_Demand" in workbook_sheets:
        ws = workbook["Stock_Demand"]
        seen_parts = {}
        count = 0
        for row_no, row in rows_until_blank(ws):
            count += 1
            values = list(row) + [None] * 4
            part, current_stock, minimum_stock, _ = values[:4]
            if _diag_value_blank(part):
                _add_diagnostic(
                    diagnostics, "ERROR", "Stock_Demand", row_no,
                    "Part Name", "Mandatory value is blank.",
                )
                continue
            part_text = str(part).strip()
            stock_parts.add(part_text)
            if part_text in seen_parts:
                _add_diagnostic(
                    diagnostics, "ERROR", "Stock_Demand", row_no,
                    "Part Name",
                    f"Duplicate part; first found at row {seen_parts[part_text]}.",
                    part_text,
                )
            else:
                seen_parts[part_text] = row_no

            for column, value in (
                ("Current FG Stock", current_stock),
                ("Minimum Stock", minimum_stock),
            ):
                number = _diag_number(value)
                if not _diag_value_blank(value) and number is None:
                    _add_diagnostic(
                        diagnostics, "ERROR", "Stock_Demand", row_no,
                        column, "Value must be numeric.", value,
                    )
                elif number is not None and number < 0:
                    _add_diagnostic(
                        diagnostics, "ERROR", "Stock_Demand", row_no,
                        column, "Value cannot be negative.", value,
                    )

        populated_counts["Stock_Demand"] = count

    # Batch configuration.
    if "Batch_Config" in workbook_sheets:
        ws = workbook["Batch_Config"]
        seen_parts = {}
        count = 0
        for row_no, row in rows_until_blank(ws):
            count += 1
            values = list(row) + [None] * 3
            part, production_batch, transportation_batch = values[:3]
            if _diag_value_blank(part):
                _add_diagnostic(
                    diagnostics, "ERROR", "Batch_Config", row_no,
                    "Part Name", "Mandatory value is blank.",
                )
                continue
            part_text = str(part).strip()
            batch_parts.add(part_text)
            if part_text in seen_parts:
                _add_diagnostic(
                    diagnostics, "ERROR", "Batch_Config", row_no,
                    "Part Name",
                    f"Duplicate part; first found at row {seen_parts[part_text]}.",
                    part_text,
                )
            else:
                seen_parts[part_text] = row_no

            for column, value in (
                ("Production Batch Qty", production_batch),
                ("Transportation Batch Qty", transportation_batch),
            ):
                number = _diag_number(value)
                if not _diag_value_blank(value) and number is None:
                    _add_diagnostic(
                        diagnostics, "ERROR", "Batch_Config", row_no,
                        column, "Value must be numeric.", value,
                    )
                elif number is not None and number < 0:
                    _add_diagnostic(
                        diagnostics, "ERROR", "Batch_Config", row_no,
                        column, "Value cannot be negative.", value,
                    )

            transport_num = _diag_number(transportation_batch)
            if transport_num == 0:
                _add_diagnostic(
                    diagnostics, "WARNING", "Batch_Config", row_no,
                    "Transportation Batch Qty",
                    "Zero means the exact process quantity will remain one lot.",
                    transportation_batch,
                )

        populated_counts["Batch_Config"] = count

    # Process BOM.
    if "Process_BOM" in workbook_sheets:
        ws = workbook["Process_BOM"]
        seen_routes = {}
        sequence_by_part = {}
        count = 0
        for row_no, row in rows_until_blank(ws):
            count += 1
            values = list(row) + [None] * 9
            (
                part, sequence, operation, process_type, cycle_time,
                setup_time, outsource_hours, multiplier, scrap
            ) = values[:9]

            for column, value in (
                ("Part Name", part),
                ("Process Sequence", sequence),
                ("Operation Name", operation),
                ("Process Type", process_type),
            ):
                if _diag_value_blank(value):
                    _add_diagnostic(
                        diagnostics, "ERROR", "Process_BOM", row_no,
                        column, "Mandatory value is blank.", value,
                    )

            if _diag_value_blank(part):
                continue
            part_text = str(part).strip()
            process_parts.add(part_text)

            seq_num = _diag_number(sequence)
            if seq_num is None:
                _add_diagnostic(
                    diagnostics, "ERROR", "Process_BOM", row_no,
                    "Process Sequence", "Sequence must be numeric.", sequence,
                )
                continue
            if seq_num <= 0 or int(seq_num) != seq_num:
                _add_diagnostic(
                    diagnostics, "ERROR", "Process_BOM", row_no,
                    "Process Sequence",
                    "Sequence must be a positive whole number.", sequence,
                )

            route_key = (
                part_text,
                int(seq_num) if seq_num is not None else sequence,
            )
            if route_key in seen_routes:
                _add_diagnostic(
                    diagnostics, "ERROR", "Process_BOM", row_no,
                    "Process Sequence",
                    f"Duplicate part/sequence; first found at row {seen_routes[route_key]}.",
                    sequence,
                )
            else:
                seen_routes[route_key] = row_no

            sequence_by_part.setdefault(part_text, []).append(
                (int(seq_num), row_no)
            )

            ptype = str(process_type or "").strip().upper()
            if ptype not in ("INHOUSE", "OUTSOURCE"):
                _add_diagnostic(
                    diagnostics, "ERROR", "Process_BOM", row_no,
                    "Process Type",
                    "Allowed values are INHOUSE or OUTSOURCE.", process_type,
                )

            if ptype == "INHOUSE":
                cycle_num = _diag_number(cycle_time)
                if cycle_num is None or cycle_num <= 0:
                    _add_diagnostic(
                        diagnostics, "ERROR", "Process_BOM", row_no,
                        "Cycle Time Sec/Part",
                        "INHOUSE operation requires cycle time greater than zero.",
                        cycle_time,
                    )
            elif ptype == "OUTSOURCE":
                lead_num = _diag_number(outsource_hours)
                if lead_num is None or lead_num <= 0:
                    _add_diagnostic(
                        diagnostics, "ERROR", "Process_BOM", row_no,
                        "Outsource Lead Time Hours",
                        "OUTSOURCE operation requires lead time greater than zero.",
                        outsource_hours,
                    )

            for column, value, default, minimum in (
                ("Setup Time Min", setup_time, 0, 0),
                ("Qty Multiplier", multiplier, 1, 0),
                ("Scrap Allowance %", scrap, 0, 0),
            ):
                if _diag_value_blank(value):
                    _add_diagnostic(
                        diagnostics, "WARNING", "Process_BOM", row_no,
                        column, f"Blank value; importer will use default {default}.",
                    )
                else:
                    number = _diag_number(value)
                    if number is None or number < minimum:
                        _add_diagnostic(
                            diagnostics, "ERROR", "Process_BOM", row_no,
                            column, "Invalid numeric value.", value,
                        )

        populated_counts["Process_BOM"] = count

    # Machine recommendations.
    if "Machine_Recommendations" in workbook_sheets:
        ws = workbook["Machine_Recommendations"]
        count = 0
        for row_no, row in rows_until_blank(ws):
            count += 1
            values = list(row)
            part = values[0] if len(values) > 0 else None
            operation = values[1] if len(values) > 1 else None
            machines = values[2:17]

            if _diag_value_blank(part):
                _add_diagnostic(
                    diagnostics, "ERROR", "Machine_Recommendations", row_no,
                    "Part Name", "Mandatory value is blank.",
                )
            if _diag_value_blank(operation):
                _add_diagnostic(
                    diagnostics, "ERROR", "Machine_Recommendations", row_no,
                    "Operation Name", "Mandatory value is blank.",
                )

            if not _diag_value_blank(part) and not _diag_value_blank(operation):
                machine_routes.add(
                    (str(part).strip(), str(operation).strip())
                )
                if not any(not _diag_value_blank(machine) for machine in machines):
                    _add_diagnostic(
                        diagnostics, "WARNING", "Machine_Recommendations",
                        row_no, "Machine 1",
                        "No recommended machine was entered.",
                    )

        populated_counts["Machine_Recommendations"] = count

    # Referential checks.
    for part in sorted(schedule_parts):
        if part not in process_parts:
            _add_diagnostic(
                diagnostics, "ERROR", "Customer_Schedules", 0, "Part Name",
                "Scheduled part has no Process_BOM route.", part,
            )
        if part not in stock_parts:
            _add_diagnostic(
                diagnostics, "ERROR", "Customer_Schedules", 0, "Part Name",
                "Scheduled part has no Stock_Demand record.", part,
            )
        if part not in batch_parts:
            _add_diagnostic(
                diagnostics, "ERROR", "Customer_Schedules", 0, "Part Name",
                "Scheduled part has no Batch_Config record.", part,
            )

    # Machine route checks for in-house BOM rows.
    if "Process_BOM" in workbook_sheets:
        ws = workbook["Process_BOM"]
        for row_no, row in rows_until_blank(ws):
            values = list(row) + [None] * 4
            part, _, operation, process_type = values[:4]
            if (
                not _diag_value_blank(part)
                and not _diag_value_blank(operation)
                and str(process_type or "").strip().upper() == "INHOUSE"
                and (str(part).strip(), str(operation).strip()) not in machine_routes
            ):
                _add_diagnostic(
                    diagnostics, "WARNING", "Process_BOM", row_no,
                    "Operation Name",
                    "No matching Machine_Recommendations row exists.",
                    operation,
                )

    # Optional calendar validation.
    if "Machine_Downtime" in workbook_sheets:
        ws = workbook["Machine_Downtime"]
        count = 0
        for row_no, row in rows_until_blank(ws):
            count += 1
            values = list(row) + [None] * 6
            machine, sd, stime, ed, etime, _ = values[:6]
            date_time_values = (sd, stime, ed, etime)
            if _diag_value_blank(machine):
                _add_diagnostic(
                    diagnostics, "WARNING", "Machine_Downtime", row_no,
                    "Machine Name", "Blank row or machine name; row will be skipped.",
                )
            elif all(_diag_value_blank(v) for v in date_time_values):
                _add_diagnostic(
                    diagnostics, "WARNING", "Machine_Downtime", row_no,
                    "Unavailable Start Date",
                    "Machine is present but downtime dates/times are blank; row will be skipped.",
                    machine,
                )
            elif any(_diag_value_blank(v) for v in date_time_values):
                _add_diagnostic(
                    diagnostics, "ERROR", "Machine_Downtime", row_no,
                    "Unavailable Date/Time",
                    "All four start/end date and time fields are required.",
                )
        populated_counts["Machine_Downtime"] = count

    if "Shifts" in workbook_sheets:
        ws = workbook["Shifts"]
        count = 0
        for row_no, row in rows_until_blank(ws):
            count += 1
            values = list(row) + [None] * 4
            shift, start_time, end_time, active = values[:4]
            if _diag_value_blank(shift):
                _add_diagnostic(
                    diagnostics, "ERROR", "Shifts", row_no,
                    "Shift Name", "Mandatory value is blank.",
                )
            if _diag_time(start_time) is None:
                _add_diagnostic(
                    diagnostics, "ERROR", "Shifts", row_no,
                    "Start Time", "Invalid or blank shift start time.", start_time,
                )
            if _diag_time(end_time) is None:
                _add_diagnostic(
                    diagnostics, "ERROR", "Shifts", row_no,
                    "End Time", "Invalid or blank shift end time.", end_time,
                )
            if str(active or "Y").strip().upper() not in ("Y", "N"):
                _add_diagnostic(
                    diagnostics, "ERROR", "Shifts", row_no,
                    "Active (Y/N)", "Allowed values are Y or N.", active,
                )
        populated_counts["Shifts"] = count

    # Build sheet summary.
    all_named_sheets = list(REQUIRED_SHEETS) + list(OPTIONAL_SHEETS)
    for sheet_name in all_named_sheets:
        if sheet_name not in workbook_sheets:
            if not any(row["Sheet"] == sheet_name for row in sheet_summary):
                sheet_summary.append({
                    "Sheet": sheet_name,
                    "Status": "Optional Missing"
                    if sheet_name in OPTIONAL_SHEETS else "Missing",
                    "Populated Rows": 0,
                })
            continue

        sheet_findings = [
            item for item in diagnostics if item["Sheet"] == sheet_name
        ]
        errors = sum(1 for item in sheet_findings if item["Severity"] == "ERROR")
        warnings = sum(
            1 for item in sheet_findings if item["Severity"] == "WARNING"
        )
        status = "Error" if errors else ("Warning" if warnings else "OK")
        sheet_summary.append({
            "Sheet": sheet_name,
            "Status": status,
            "Populated Rows": populated_counts.get(sheet_name, 0),
            "Errors": errors,
            "Warnings": warnings,
        })

    workbook.close()

    error_count = sum(
        1 for item in diagnostics if item["Severity"] == "ERROR"
    )
    warning_count = sum(
        1 for item in diagnostics if item["Severity"] == "WARNING"
    )

    return {
        "valid_for_import": error_count == 0,
        "error_count": error_count,
        "warning_count": warning_count,
        "diagnostics": diagnostics,
        "sheet_summary": sheet_summary,
    }


def uploaded_file_signature(uploaded_file):
    raw = uploaded_file.getvalue()
    return hashlib.sha256(raw).hexdigest()


st.title("ELECTRO-DIP")
st.subheader("Online Production Planning System")
st.caption(
    "Production planning, operator slips, persistent entries, "
    "transportation lots, operation sequence control and WIP reports."
)

with st.container(border=True):
    st.markdown("### Upload, Validate & Import Excel")
    upload_col, template_col = st.columns([2.2, 1])

    with upload_col:
        uploaded_excel = st.file_uploader(
            "Select Production Planning Excel",
            type=["xlsx", "xlsm"],
            key="master_excel_upload",
        )

    with template_col:
        if TEMPLATE_PATH.exists():
            st.download_button(
                "Download Import Template",
                data=TEMPLATE_PATH.read_bytes(),
                file_name=TEMPLATE_PATH.name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

    if uploaded_excel is not None:
        current_signature = uploaded_file_signature(uploaded_excel)
        st.caption(
            f"Selected file: {uploaded_excel.name} "
            f"({uploaded_excel.size / 1024:.1f} KB)"
        )

        button_col1, button_col2 = st.columns(2)

        with button_col1:
            validate_clicked = st.button(
                "1. Validate Template",
                type="secondary",
                use_container_width=True,
            )

        if validate_clicked:
            try:
                validation_result = validate_import_template(uploaded_excel)
                st.session_state["template_validation"] = validation_result
                st.session_state["template_validation_signature"] = (
                    current_signature
                )
            except Exception as exc:
                st.session_state.pop("template_validation", None)
                st.session_state.pop(
                    "template_validation_signature", None
                )
                st.error(f"Template validation failed: {exc}")

        validation = st.session_state.get("template_validation")
        validated_signature = st.session_state.get(
            "template_validation_signature"
        )
        validation_current = (
            validation is not None
            and validated_signature == current_signature
        )

        with button_col2:
            import_clicked = st.button(
                "2. Import Excel",
                type="primary",
                disabled=(
                    not validation_current
                    or not validation.get("valid_for_import", False)
                ),
                use_container_width=True,
            )

        if validation_current:
            metric1, metric2, metric3 = st.columns(3)
            metric1.metric(
                "Validation Status",
                "PASS" if validation["valid_for_import"] else "FAILED",
            )
            metric2.metric("Errors", validation["error_count"])
            metric3.metric("Warnings", validation["warning_count"])

            summary_df = pd.DataFrame(validation["sheet_summary"])
            st.markdown("#### Worksheet Health Check")
            st.dataframe(
                summary_df,
                hide_index=True,
                use_container_width=True,
            )

            diagnostics_df = pd.DataFrame(validation["diagnostics"])
            if diagnostics_df.empty:
                st.success(
                    "No validation errors or warnings. The workbook is ready "
                    "for import."
                )
            else:
                st.markdown("#### Detailed Diagnostics")
                severity_filter = st.multiselect(
                    "Severity",
                    ["ERROR", "WARNING"],
                    default=["ERROR", "WARNING"],
                    key="diagnostic_severity_filter",
                )
                filtered_diagnostics = diagnostics_df[
                    diagnostics_df["Severity"].isin(severity_filter)
                ]
                st.dataframe(
                    filtered_diagnostics,
                    hide_index=True,
                    use_container_width=True,
                    height=360,
                )

                csv_data = diagnostics_df.to_csv(
                    index=False
                ).encode("utf-8")
                st.download_button(
                    "Download Diagnostic Report CSV",
                    data=csv_data,
                    file_name="Electro_Dip_Import_Diagnostics.csv",
                    mime="text/csv",
                )

            if not validation["valid_for_import"]:
                st.error(
                    "Import is blocked until all ERROR items are corrected. "
                    "WARNING items do not block import."
                )
            elif validation["warning_count"] > 0:
                st.warning(
                    "Validation passed with warnings. Review them before "
                    "importing."
                )

        elif validation is not None:
            st.warning(
                "The selected file changed after validation. Click "
                "'Validate Template' again."
            )

        if import_clicked:
            try:
                backup_info = db.create_database_backup(
                    reason="BEFORE_IMPORT"
                )
                uploaded_excel.seek(0)
                import_report = db.import_workbook(uploaded_excel)
                import_report["backup_created"] = backup_info
                st.session_state["latest_import_report"] = import_report
                st.success(
                    "Excel imported successfully. "
                    f"Previous entries preserved: "
                    f"{import_report['previous_entries_preserved']}"
                )
                st.rerun()
            except Exception as exc:
                st.error(f"Excel import failed: {exc}")
    else:
        st.info(
            "Select an Excel workbook, run Validate Template, correct any "
            "errors, and then import."
        )

    if "latest_import_report" in st.session_state:
        with st.expander("Latest Import Report"):
            latest_report = st.session_state["latest_import_report"]
            counts = latest_report.get("counts", {})
            if counts:
                st.dataframe(
                    pd.DataFrame(
                        [
                            {"Data Type": key, "Imported Rows": value}
                            for key, value in counts.items()
                        ]
                    ),
                    hide_index=True,
                    use_container_width=True,
                )
            warnings = latest_report.get("warnings", [])
            if warnings:
                for warning in warnings[:100]:
                    st.write("•", warning)
            else:
                st.success("No import warnings.")

tabs = st.tabs([
    "Dashboard", "Production Plan", "Operator Slips", "Operator Entry",
    "Process WIP Report", "WIP Ageing", "WIP Summaries",
    "Schedule Calculation", "Process WIP Allocation",
    "Previous Entries", "Download Excel", "Clear Data"
])

with tabs[0]:
    live_ist = datetime.now(ZoneInfo("Asia/Kolkata"))
    st.info(
        f"Live India Time: {live_ist:%d-%b-%Y %H:%M:%S} IST"
    )
    entries = db.conn.execute("SELECT COUNT(*) FROM operator_entries").fetchone()[0]
    wip_total = sum(r["wip_after_process"] for r in db.wip_rows())
    dispatched = db.conn.execute(
        """SELECT COALESCE(SUM(e.good_qty),0)
           FROM operator_entries e
           WHERE e.process_sequence = (
               SELECT MAX(p.process_sequence)
               FROM process_bom p WHERE p.part_name=e.part_name
           )"""
    ).fetchone()[0]
    plan_rows = db.conn.execute("SELECT COUNT(*) FROM production_plan").fetchone()[0]
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Saved Entries", entries)
    c2.metric("Physical WIP Qty", round(wip_total, 2))
    c3.metric("Dispatched Good Qty", round(float(dispatched or 0), 2))
    c4.metric("Plan Rows", plan_rows)
    st.info(
        "Highest Process BOM sequence is treated as Dispatch. "
        "Dispatch quantity is allocated FIFO to the earliest due schedule. "
        "Minimum stock is generated once, after customer schedules."
    )


    with st.expander("Database Health & Backups"):
        health = db.database_health()
        health_df = pd.DataFrame(
            [{"Check": key, "Value": value} for key, value in health.items()]
        )
        st.dataframe(health_df, hide_index=True, use_container_width=True)

        manual_backup_col, history_col = st.columns([1, 2])
        with manual_backup_col:
            if st.button("Create Manual Backup", use_container_width=True):
                backup = db.create_database_backup("MANUAL")
                st.success(
                    f"Backup created: {backup['filename']} "
                    f"at {backup['timestamp']} IST"
                )
        with history_col:
            backup_history = pd.DataFrame(db.backup_history_rows())
            if not backup_history.empty:
                st.dataframe(
                    backup_history,
                    hide_index=True,
                    use_container_width=True,
                    height=220,
                )

with tabs[1]:
    st.subheader("Production Plan")
    st.info(
        "V14 schedule quantity remains unchanged. Transportation batch only "
        "splits the exact process quantity into lot rows. Example: "
        "240 with batch 100 becomes 100 + 100 + 40."
    )
    p1, p2 = st.columns([1, 1])
    with p1:
        if st.button("Generate / Regenerate Production Plan", type="primary"):
            count = db.generate_plan()
            st.success(f"Production plan generated: {count} rows")
    with p2:
        plan_rows_for_download = db.production_plan_report_rows()
        if plan_rows_for_download:
            st.download_button(
                "Download Production Plan Excel",
                data=make_production_plan_excel_bytes(),
                file_name="Electro_Dip_Production_Plan.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

    plan_df = pd.DataFrame(db.production_plan_report_rows())
    if plan_df.empty:
        st.info(
            "No production plan is available. Import the Excel template and "
            "click Generate / Regenerate Production Plan."
        )
    else:
        preferred_plan_columns = [
            "plan_id", "schedule_id", "customer_name", "part_name",
            "process_sequence", "operation_name", "machine_name",
            "shift_name", "schedule_plan_qty_v14",
            "net_process_plan_qty", "transportation_batch_qty",
            "lot_no", "lot_planned_qty",
            "production_start_datetime", "production_end_datetime",
            "due_datetime", "priority", "note",
        ]
        visible_plan_columns = [
            column for column in preferred_plan_columns
            if column in plan_df.columns
        ]
        st.dataframe(
            plan_df[visible_plan_columns],
            hide_index=True,
            use_container_width=True,
            height=570,
        )
        st.caption(
            "Schedule Plan Qty (V14) is the approved schedule calculation. "
            "Net Process Plan Qty is after process-wise WIP. Lot Planned Qty "
            "is the transportation-batch split; all lot totals equal the exact "
            "net process quantity."
        )

with tabs[2]:
    st.subheader("Machine-wise Operator Slips")
    slip_dates = db.operator_slip_dates()
    if not slip_dates:
        st.info("Generate the production plan before creating operator slips.")
    else:
        s1, s2 = st.columns(2)
        selected_slip_date = s1.selectbox(
            "Production Date",
            slip_dates,
            format_func=lambda d: d.strftime("%d-%b-%Y"),
        )
        machines = ["All Machines"] + db.operator_slip_machines(selected_slip_date)
        selected_machine_label = s2.selectbox("Machine", machines)
        selected_machine = (
            None if selected_machine_label == "All Machines"
            else selected_machine_label
        )

        slip_rows = db.operator_slip_rows(
            due_date=selected_slip_date,
            machine_name=selected_machine,
        )
        slip_df = pd.DataFrame(slip_rows)
        if not slip_df.empty:
            preferred = [
                "plan_id", "machine_name", "customer_name", "part_name",
                "process_sequence", "operation_name", "planned_qty",
                "production_start_datetime", "production_end_datetime",
                "shift_name", "lot_no", "schedule_id",
                "due_datetime", "priority"
            ]
            visible = [c for c in preferred if c in slip_df.columns]
            st.dataframe(
                slip_df[visible],
                hide_index=True,
                use_container_width=True,
                height=440,
            )
            st.download_button(
                "Download / Print Operator Slip PDF",
                data=make_operator_slip_pdf_bytes(
                    selected_slip_date,
                    selected_machine,
                ),
                file_name=(
                    f"Operator_Slips_{selected_slip_date.isoformat()}.pdf"
                ),
                mime="application/pdf",
                type="primary",
            )
            st.caption(
                "One page is generated per machine. Multiple operations are "
                "printed on the same approved ELECTRO-DIP slip."
            )

with tabs[3]:
    st.subheader("Daily Operator Entry")

    plan_rows = db.plan_id_rows()
    if not plan_rows:
        st.warning("Generate the production plan first.")
    else:
        plan_ids = [row["plan_id"] for row in plan_rows]

        top1, top2 = st.columns([2, 1])
        selected_plan_id = top1.selectbox(
            "Plan ID",
            plan_ids,
            key="operator_plan_id",
        )
        plan_detail = db.plan_id_detail(selected_plan_id)
        gate_status = db.sequence_gate_status(selected_plan_id)

        with top2:
            st.caption("Selected plan")
            st.write(
                f"{plan_detail['part_name']} | "
                f"{plan_detail['operation_name']} | "
                f"{plan_detail.get('machine_name') or '-'} | "
                f"{plan_detail.get('shift_name') or '-'} | "
                f"Qty {plan_detail['planned_qty']:g}"
            )

        part = str(plan_detail["part_name"])
        sequence = int(plan_detail["process_sequence"])
        operation = str(plan_detail["operation_name"])
        customer_name = str(plan_detail.get("customer_name") or "")
        schedule_id = str(plan_detail.get("schedule_id") or "")
        planned_default = float(plan_detail.get("planned_qty") or 0)

        production_start = plan_detail.get("production_start_datetime")
        production_date = (
            pd.to_datetime(production_start).date()
            if production_start else date.today()
        )

        machine = str(plan_detail.get("machine_name") or "")
        shift = str(plan_detail.get("shift_name") or "")
        production_end = plan_detail.get("production_end_datetime")

        start_time_text = (
            pd.to_datetime(production_start).strftime("%H:%M")
            if production_start else ""
        )
        end_time_text = (
            pd.to_datetime(production_end).strftime("%H:%M")
            if production_end else ""
        )

        c1, c2 = st.columns(2)
        c1.text_input("Part Number", value=part, disabled=True)
        c2.text_input(
            "Process / Sequence",
            value=f"{sequence} - {operation}",
            disabled=True,
        )

        c3, c4, c5 = st.columns(3)
        c3.text_input(
            "Production Date",
            value=production_date.strftime("%Y/%m/%d"),
            disabled=True,
        )
        c4.text_input("Machine", value=machine, disabled=True)
        c5.text_input("Shift", value=shift, disabled=True)

        c6, c7, c8 = st.columns(3)
        c6.text_input(
            "Production Start Time",
            value=start_time_text,
            disabled=True,
        )
        c7.text_input(
            "Production End Time",
            value=end_time_text,
            disabled=True,
        )
        c8.text_input(
            "Planned Qty",
            value=f"{planned_default:g}",
            disabled=True,
        )

        st.markdown("#### Operation Sequence Control")
        g1, g2, g3, g4 = st.columns(4)
        g1.metric("Lot Planned Qty", f"{gate_status['lot_planned_qty']:g}")
        g2.metric(
            "Previous Operation Good Qty",
            f"{gate_status['previous_good_qty']:g}",
        )
        g3.metric(
            "Already Processed Here",
            f"{gate_status['already_processed_here']:g}",
        )
        g4.metric(
            "Maximum Entry Allowed",
            f"{gate_status['maximum_entry_allowed']:g}",
        )

        if gate_status["gate_open"]:
            st.success(gate_status["reason"])
        else:
            st.error(gate_status["reason"])

        c9, c10 = st.columns(2)
        actual = c9.number_input(
            "Actual Qty",
            min_value=0.0,
            max_value=float(gate_status["maximum_entry_allowed"]),
            step=1.0,
            disabled=not gate_status["gate_open"],
            key=f"actual_{selected_plan_id}",
        )
        rejected = c10.number_input(
            "Rejected Qty",
            min_value=0.0,
            max_value=float(actual),
            step=1.0,
            disabled=not gate_status["gate_open"],
            key=f"rejected_{selected_plan_id}",
        )

        entry_date = production_date
        planned = planned_default

        operator_names = db.personnel_names("OPERATOR")
        supervisor_names = db.personnel_names("SUPERVISOR")

        p1, p2 = st.columns(2)
        operator = p1.selectbox(
            "Operator Name",
            operator_names if operator_names else [""],
            key=f"operator_{selected_plan_id}",
        )
        supervisor = p2.selectbox(
            "Supervisor Name",
            supervisor_names if supervisor_names else [""],
            key=f"supervisor_{selected_plan_id}",
        )

        with st.expander("Add Operator / Supervisor Name"):
            a1, a2, a3 = st.columns([2, 1, 1])
            new_person_name = a1.text_input("Name")
            new_person_role = a2.selectbox(
                "Role", ["OPERATOR", "SUPERVISOR"]
            )
            if a3.button("Add Name"):
                try:
                    db.add_personnel(new_person_name, new_person_role)
                    st.success("Name added. Refresh this tab to see it in the dropdown.")
                except Exception as exc:
                    st.error(str(exc))

        status = st.selectbox(
            "Status",
            ["Running", "Completed", "Hold", "Rework"],
            key=f"status_{selected_plan_id}",
        )
        remarks = st.text_area(
            "Remarks",
            key=f"remarks_{selected_plan_id}",
        )

        if st.button(
            "Save Operator Entry",
            type="primary",
            disabled=not gate_status["gate_open"],
        ):
            if rejected > actual:
                st.error("Rejected Qty cannot exceed Actual Qty.")
            elif not operator:
                st.error("Select an Operator Name.")
            elif not supervisor:
                st.error("Select a Supervisor Name.")
            else:
                try:
                    good = db.add_operator_entry(
                        entry_date=entry_date,
                        part_name=part,
                        process_sequence=sequence,
                        operation_name=operation,
                        actual_qty=actual,
                        rejected_qty=rejected,
                        machine_name=machine,
                        shift_name=shift,
                        plan_id=selected_plan_id,
                        schedule_id=schedule_id,
                        customer_name=customer_name,
                        planned_qty=planned,
                        status=status,
                        operator_name=operator,
                        supervisor_name=supervisor,
                        remarks=remarks,
                    )
                    st.success(
                        f"Entry saved for Plan ID {selected_plan_id}. "
                        f"Good Qty = {good:g}."
                    )
                except Exception as exc:
                    st.error(str(exc))

        today_entries = pd.DataFrame([
            dict(r) for r in db.operator_entries()
            if str(r["entry_date"]) == entry_date.isoformat()
        ])
        st.markdown(f"#### Entries for Production Date ({entry_date:%d-%b-%Y})")
        if today_entries.empty:
            st.info("No entries for the selected date.")
        else:
            preferred = [
                "plan_id", "part_name", "process_sequence",
                "operation_name", "machine_name", "shift_name",
                "planned_qty", "actual_qty", "rejected_qty",
                "operator_name", "supervisor_name", "status",
                "live_entry_date", "live_entry_time",
                "entry_timestamp", "last_modified_timestamp",
            ]
            visible = [c for c in preferred if c in today_entries.columns]
            st.dataframe(
                today_entries[visible],
                hide_index=True,
                use_container_width=True,
            )

with tabs[4]:
    st.subheader("Process-wise WIP Report")
    report_df = pd.DataFrame(db.process_wip_report_rows())
    if report_df.empty:
        st.info("No WIP data.")
    else:
        f1, f2, f3, f4 = st.columns(4)
        customers = ["All"] + sorted([x for x in report_df["customer_name"].dropna().astype(str).unique() if x])
        parts = ["All"] + sorted(report_df["part_name"].astype(str).unique())
        processes = ["All"] + sorted(report_df["operation_name"].astype(str).unique())
        machines = ["All"] + sorted([x for x in report_df["machine_name"].dropna().astype(str).unique() if x])
        customer = f1.selectbox("Customer", customers)
        part = f2.selectbox("Part", parts)
        process = f3.selectbox("Process", processes)
        machine = f4.selectbox("Machine", machines)

        filtered = report_df.copy()
        if customer != "All":
            filtered = filtered[filtered["customer_name"].astype(str) == customer]
        if part != "All":
            filtered = filtered[filtered["part_name"].astype(str) == part]
        if process != "All":
            filtered = filtered[filtered["operation_name"].astype(str) == process]
        if machine != "All":
            filtered = filtered[filtered["machine_name"].astype(str) == machine]

        st.dataframe(filtered, hide_index=True, use_container_width=True, height=540)
        st.caption("WIP Qty = normalized cumulative good at current process minus normalized cumulative good at next process.")

with tabs[5]:
    st.subheader("WIP Ageing")
    ageing_df = pd.DataFrame(db.process_wip_report_rows())
    ageing_df = ageing_df[ageing_df["wip_qty"] > 0] if not ageing_df.empty else ageing_df
    if ageing_df.empty:
        st.info("No open WIP.")
    else:
        a1, a2, a3, a4 = st.columns(4)
        a1.metric("Total WIP", round(float(ageing_df["wip_qty"].sum()), 2))
        a2.metric("Fresh", round(float(ageing_df.loc[ageing_df["age_status"]=="Fresh", "wip_qty"].sum()), 2))
        a3.metric("Monitor", round(float(ageing_df.loc[ageing_df["age_status"]=="Monitor", "wip_qty"].sum()), 2))
        a4.metric("Old WIP", round(float(ageing_df.loc[ageing_df["age_status"]=="Old WIP", "wip_qty"].sum()), 2))
        st.dataframe(
            ageing_df.sort_values(["wip_age_days", "wip_qty"], ascending=[False, False]),
            hide_index=True, use_container_width=True, height=520
        )

with tabs[6]:
    st.subheader("Machine / Part / Customer WIP")
    s1, s2, s3 = st.columns(3)
    with s1:
        st.markdown("#### Machine-wise")
        st.dataframe(pd.DataFrame(db.machine_wip_rows()), hide_index=True, use_container_width=True)
    with s2:
        st.markdown("#### Part-wise")
        st.dataframe(pd.DataFrame(db.part_wip_rows()), hide_index=True, use_container_width=True)
    with s3:
        st.markdown("#### Customer-wise")
        st.dataframe(pd.DataFrame(db.customer_wip_rows()), hide_index=True, use_container_width=True)

with tabs[7]:
    st.subheader("Schedule-Line Plan Quantity Calculation")
    schedule_calc_df = pd.DataFrame(db.schedule_line_calculation_rows())
    if schedule_calc_df.empty:
        st.info("No schedules imported.")
    else:
        st.dataframe(schedule_calc_df, hide_index=True, use_container_width=True, height=520)
        st.caption("First line: Demand + Minimum Stock - Opening Available. Further lines: Demand + Minimum Stock - Previous Schedule Carry-Forward WIP.")

with tabs[8]:
    st.subheader("Process-BOM-Wise WIP Allocation")
    process_wip_df = pd.DataFrame(db.process_schedule_wip_rows())
    if process_wip_df.empty:
        st.info("No process WIP allocation available.")
    else:
        st.dataframe(process_wip_df, hide_index=True, use_container_width=True, height=520)

with tabs[9]:
    st.subheader("Previous Operator Entries")
    st.caption(
        "Use Edit to correct Actual Qty, Rejected Qty, names, status or remarks. "
        "Use Delete only for a completely wrong entry. WIP and sequence control "
        "recalculate automatically."
    )

    entries = [dict(r) for r in db.operator_entries()]
    entries_df = pd.DataFrame(entries)

    if not entries_df.empty:
        previous_export = entries_df.to_excel(
            index=False, engine="openpyxl"
        ) if False else None
        csv_bytes = entries_df.to_csv(index=False).encode("utf-8")
        st.download_button(
            "Download Previous Entries CSV",
            data=csv_bytes,
            file_name="Electro_Dip_Previous_Entries.csv",
            mime="text/csv",
        )

    if entries_df.empty:
        st.info("No saved operator entries.")
    else:
        filter_col1, filter_col2, filter_col3 = st.columns(3)

        entry_dates = ["All"] + sorted(
            entries_df["entry_date"].dropna().astype(str).unique().tolist(),
            reverse=True,
        )
        parts = ["All"] + sorted(
            entries_df["part_name"].dropna().astype(str).unique().tolist()
        )
        plan_ids = ["All"] + sorted(
            [
                value for value in
                entries_df["plan_id"].dropna().astype(str).unique().tolist()
                if value
            ]
        )

        selected_date = filter_col1.selectbox(
            "Entry Date",
            entry_dates,
            key="previous_entry_date_filter",
        )
        selected_part = filter_col2.selectbox(
            "Part Number",
            parts,
            key="previous_entry_part_filter",
        )
        selected_plan = filter_col3.selectbox(
            "Plan ID",
            plan_ids,
            key="previous_entry_plan_filter",
        )

        filtered_entries = entries_df.copy()
        if selected_date != "All":
            filtered_entries = filtered_entries[
                filtered_entries["entry_date"].astype(str) == selected_date
            ]
        if selected_part != "All":
            filtered_entries = filtered_entries[
                filtered_entries["part_name"].astype(str) == selected_part
            ]
        if selected_plan != "All":
            filtered_entries = filtered_entries[
                filtered_entries["plan_id"].astype(str) == selected_plan
            ]

        display_columns = [
            "entry_id", "entry_date", "plan_id", "schedule_id",
            "customer_name", "part_name", "process_sequence",
            "operation_name", "machine_name", "shift_name",
            "planned_qty", "actual_qty", "rejected_qty", "good_qty",
            "operator_name", "supervisor_name", "status",
            "remarks", "live_entry_date", "live_entry_time",
            "entry_timestamp", "last_modified_timestamp",
        ]
        visible_columns = [
            column for column in display_columns
            if column in filtered_entries.columns
        ]

        st.dataframe(
            filtered_entries[visible_columns],
            hide_index=True,
            use_container_width=True,
            height=430,
        )

        st.markdown("### Edit or Delete Entry")
        available_ids = filtered_entries["entry_id"].astype(int).tolist()

        if not available_ids:
            st.info("No entries match the selected filters.")
        else:
            selected_entry_id = st.selectbox(
                "Select Entry ID",
                available_ids,
                key="selected_previous_entry_id",
            )

            selected_entry = db.operator_entry_by_id(selected_entry_id)

            if selected_entry:
                st.info(
                    f"Plan ID: {selected_entry.get('plan_id') or '-'} | "
                    f"Part: {selected_entry['part_name']} | "
                    f"Operation: {selected_entry['process_sequence']} - "
                    f"{selected_entry['operation_name']} | "
                    f"Machine: {selected_entry.get('machine_name') or '-'} | "
                    f"Live Entry: {selected_entry.get('entry_timestamp') or '-'} IST | "
                    f"Last Modified: "
                    f"{selected_entry.get('last_modified_timestamp') or 'Never'}"
                )

                edit_col1, edit_col2 = st.columns(2)

                edit_actual = edit_col1.number_input(
                    "Actual Qty",
                    min_value=0.0,
                    value=float(selected_entry["actual_qty"] or 0),
                    step=1.0,
                    key=f"edit_actual_{selected_entry_id}",
                )
                edit_rejected = edit_col2.number_input(
                    "Rejected Qty",
                    min_value=0.0,
                    value=float(selected_entry["rejected_qty"] or 0),
                    step=1.0,
                    key=f"edit_rejected_{selected_entry_id}",
                )

                operator_names = db.personnel_names("OPERATOR")
                supervisor_names = db.personnel_names("SUPERVISOR")

                current_operator = str(
                    selected_entry.get("operator_name") or ""
                )
                current_supervisor = str(
                    selected_entry.get("supervisor_name") or ""
                )

                if current_operator and current_operator not in operator_names:
                    operator_names = [current_operator] + operator_names
                if current_supervisor and current_supervisor not in supervisor_names:
                    supervisor_names = [current_supervisor] + supervisor_names

                name_col1, name_col2 = st.columns(2)
                edit_operator = name_col1.selectbox(
                    "Operator Name",
                    operator_names if operator_names else [""],
                    index=(
                        operator_names.index(current_operator)
                        if current_operator in operator_names else 0
                    ),
                    key=f"edit_operator_{selected_entry_id}",
                )
                edit_supervisor = name_col2.selectbox(
                    "Supervisor Name",
                    supervisor_names if supervisor_names else [""],
                    index=(
                        supervisor_names.index(current_supervisor)
                        if current_supervisor in supervisor_names else 0
                    ),
                    key=f"edit_supervisor_{selected_entry_id}",
                )

                status_options = [
                    "Running", "Completed", "Hold", "Rework"
                ]
                current_status = str(selected_entry.get("status") or "")
                if current_status and current_status not in status_options:
                    status_options = [current_status] + status_options

                edit_status = st.selectbox(
                    "Status",
                    status_options,
                    index=(
                        status_options.index(current_status)
                        if current_status in status_options else 0
                    ),
                    key=f"edit_status_{selected_entry_id}",
                )

                edit_remarks = st.text_area(
                    "Remarks",
                    value=str(selected_entry.get("remarks") or ""),
                    key=f"edit_remarks_{selected_entry_id}",
                )

                action_col1, action_col2 = st.columns(2)

                with action_col1:
                    if st.button(
                        "Save Changes",
                        type="primary",
                        use_container_width=True,
                    ):
                        try:
                            good_qty = db.update_operator_entry(
                                entry_id=selected_entry_id,
                                actual_qty=edit_actual,
                                rejected_qty=edit_rejected,
                                status=edit_status,
                                operator_name=edit_operator,
                                supervisor_name=edit_supervisor,
                                remarks=edit_remarks,
                            )
                            st.success(
                                f"Entry {selected_entry_id} updated. "
                                f"Good Qty = {good_qty:g}. "
                                "WIP recalculated automatically."
                            )
                            st.rerun()
                        except Exception as exc:
                            st.error(str(exc))

                with action_col2:
                    delete_confirm = st.checkbox(
                        "Confirm permanent deletion",
                        key=f"delete_confirm_{selected_entry_id}",
                    )
                    if st.button(
                        "Delete Entry",
                        type="secondary",
                        disabled=not delete_confirm,
                        use_container_width=True,
                    ):
                        try:
                            deleted = db.delete_operator_entry(
                                selected_entry_id
                            )
                            st.success(
                                f"Entry {selected_entry_id} deleted. "
                                "WIP and sequence controls recalculated."
                            )
                            st.rerun()
                        except Exception as exc:
                            st.error(str(exc))

with tabs[10]:
    st.subheader("Download Complete WIP Workbook")
    if st.button("Prepare Downloadable Excel"):
        export_path = BASE_DIR / "Electro_Dip_WIP_Report.xlsx"
        db.export_wip_excel(export_path)
        st.session_state["wip_export"] = export_path.read_bytes()
    if "wip_export" in st.session_state:
        st.download_button(
            "Download WIP Excel",
            st.session_state["wip_export"],
            file_name="Electro_Dip_WIP_Report.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
        )
    st.write(
        "Workbook sheets include Process WIP Report, WIP Ageing data, Machine WIP, Part WIP, Customer WIP, Dispatch History, Schedule Calculation, Production Plan and Operator Entries."
    )

with tabs[11]:
    st.subheader("Clear Previous Entries")
    st.error(
        "This is the only action that deletes saved operator entries and WIP history."
    )
    confirm = st.text_input('Type exactly: CLEAR ALL ENTRIES')
    if st.button("Clear Previous Entries and Plan", type="primary"):
        if confirm != "CLEAR ALL ENTRIES":
            st.warning("Confirmation text does not match.")
        else:
            backup = db.create_database_backup("BEFORE_CLEAR_ALL")
            count = db.clear_previous_entries()
            st.session_state.pop("wip_export", None)
            st.success(
                f"{count} previous entries cleared. Production plan also "
                f"cleared. Backup created: {backup['filename']}."
            )
