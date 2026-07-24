
import io
import math
import re
import sqlite3
import shutil
from collections import defaultdict
from datetime import date, datetime, time, timedelta
from pathlib import Path
from zoneinfo import ZoneInfo

import pandas as pd
from openpyxl import load_workbook
import xlsxwriter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas
from reportlab.graphics.barcode import code128



IST_ZONE = ZoneInfo("Asia/Kolkata")


def india_now():
    """Current timezone-aware India Standard Time."""
    return datetime.now(IST_ZONE)


def india_timestamp_text():
    return india_now().strftime("%Y-%m-%d %H:%M:%S")


def is_blank(value):
    return value is None or str(value).strip() == ""


def parse_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        return (datetime(1899, 12, 30) + timedelta(days=float(value))).date()
    text = str(value).strip()
    for fmt in (
        "%Y-%m-%d",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%dT%H:%M:%S",
        "%d-%b-%Y",
        "%d/%m/%Y",
        "%d-%m-%Y",
    ):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    raise ValueError(f"Invalid date: {value}")


def parse_time(value):
    if isinstance(value, datetime):
        return value.time().replace(second=0, microsecond=0)
    if isinstance(value, time):
        return value.replace(second=0, microsecond=0)
    if isinstance(value, (int, float)):
        seconds = round((float(value) % 1) * 86400) % 86400
        return time(seconds // 3600, (seconds % 3600) // 60)
    text = str(value).strip()
    for fmt in ("%H:%M", "%H:%M:%S", "%I:%M %p"):
        try:
            return datetime.strptime(text, fmt).time()
        except ValueError:
            pass
    raise ValueError(f"Invalid time: {value}")


def combine_date_time(d, t):
    return datetime.combine(parse_date(d), parse_time(t))


class Database:
    def __init__(self, path):
        self.path = Path(path)
        self.conn = sqlite3.connect(self.path, timeout=30)
        self.conn.row_factory = sqlite3.Row
        self.create_schema()

    def close(self):
        self.conn.close()

    def create_schema(self):
        self.conn.executescript("""
        PRAGMA journal_mode=WAL;

        CREATE TABLE IF NOT EXISTS customer_schedules (
            schedule_id TEXT PRIMARY KEY,
            customer_name TEXT NOT NULL,
            part_name TEXT NOT NULL,
            customer_qty REAL NOT NULL,
            due_datetime TEXT NOT NULL,
            priority INTEGER NOT NULL DEFAULT 999
        );

        CREATE TABLE IF NOT EXISTS stock_demand (
            part_name TEXT PRIMARY KEY,
            current_stock REAL NOT NULL DEFAULT 0,
            minimum_stock REAL NOT NULL DEFAULT 0,
            remarks TEXT
        );

        CREATE TABLE IF NOT EXISTS batch_config (
            part_name TEXT PRIMARY KEY,
            production_batch REAL NOT NULL DEFAULT 0,
            transportation_batch REAL NOT NULL DEFAULT 0
        );

        CREATE TABLE IF NOT EXISTS process_bom (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            part_name TEXT NOT NULL,
            process_sequence INTEGER NOT NULL,
            operation_name TEXT NOT NULL,
            process_type TEXT NOT NULL,
            cycle_time_sec REAL NOT NULL DEFAULT 0,
            setup_time_min REAL NOT NULL DEFAULT 0,
            outsource_lead_hours REAL NOT NULL DEFAULT 0,
            qty_multiplier REAL NOT NULL DEFAULT 1,
            scrap_allowance REAL NOT NULL DEFAULT 0
        );

        CREATE TABLE IF NOT EXISTS machine_recommendations (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            part_name TEXT NOT NULL,
            operation_name TEXT NOT NULL,
            machine_name TEXT NOT NULL,
            preference_order INTEGER NOT NULL
        );

        CREATE TABLE IF NOT EXISTS machine_downtime (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            machine_name TEXT NOT NULL,
            start_datetime TEXT NOT NULL,
            end_datetime TEXT NOT NULL,
            reason TEXT DEFAULT ''
        );

        CREATE TABLE IF NOT EXISTS shifts (
            shift_name TEXT PRIMARY KEY,
            start_time TEXT NOT NULL,
            end_time TEXT NOT NULL,
            active INTEGER NOT NULL DEFAULT 1
        );

        CREATE TABLE IF NOT EXISTS breaks (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            shift_name TEXT NOT NULL,
            break_name TEXT NOT NULL,
            start_time TEXT NOT NULL,
            end_time TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS holidays (
            holiday_date TEXT PRIMARY KEY,
            holiday_name TEXT DEFAULT ''
        );

        CREATE TABLE IF NOT EXISTS weekly_offs (
            day_number INTEGER PRIMARY KEY,
            day_name TEXT NOT NULL,
            is_off INTEGER NOT NULL DEFAULT 0
        );

        CREATE TABLE IF NOT EXISTS personnel_master (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            person_name TEXT NOT NULL,
            role_name TEXT NOT NULL,
            active INTEGER NOT NULL DEFAULT 1,
            UNIQUE(person_name, role_name)
        );

        CREATE TABLE IF NOT EXISTS operator_entries (
            entry_id INTEGER PRIMARY KEY AUTOINCREMENT,
            plan_id TEXT DEFAULT '',
            entry_date TEXT NOT NULL,
            schedule_id TEXT DEFAULT '',
            customer_name TEXT DEFAULT '',
            part_name TEXT NOT NULL,
            process_sequence INTEGER NOT NULL,
            operation_name TEXT NOT NULL,
            machine_name TEXT DEFAULT '',
            shift_name TEXT DEFAULT '',
            planned_qty REAL NOT NULL DEFAULT 0,
            actual_qty REAL NOT NULL DEFAULT 0,
            rejected_qty REAL NOT NULL DEFAULT 0,
            good_qty REAL NOT NULL DEFAULT 0,
            status TEXT DEFAULT '',
            operator_name TEXT DEFAULT '',
            supervisor_name TEXT DEFAULT '',
            remarks TEXT DEFAULT '',
            source TEXT DEFAULT 'MANUAL',
            live_entry_date TEXT,
            live_entry_time TEXT,
            entry_timestamp TEXT,
            last_modified_timestamp TEXT,
            created_at TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS backup_history (
            backup_id INTEGER PRIMARY KEY AUTOINCREMENT,
            backup_timestamp TEXT NOT NULL,
            backup_reason TEXT NOT NULL,
            backup_filename TEXT NOT NULL,
            backup_size_bytes INTEGER NOT NULL DEFAULT 0
        );

        CREATE TABLE IF NOT EXISTS production_plan (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            plan_id TEXT NOT NULL,
            requirement_type TEXT NOT NULL,
            schedule_id TEXT DEFAULT '',
            customer_name TEXT DEFAULT '',
            part_name TEXT NOT NULL,
            process_sequence INTEGER NOT NULL,
            operation_name TEXT NOT NULL,
            machine_name TEXT DEFAULT '',
            shift_name TEXT DEFAULT '',
            planned_qty REAL NOT NULL,
            lot_no INTEGER NOT NULL,
            due_datetime TEXT,
            production_start_datetime TEXT,
            production_end_datetime TEXT,
            note TEXT DEFAULT ''
        );
        """)
        # Migrate older databases without deleting saved operator entries.
        existing_columns = {
            row[1] for row in self.conn.execute(
                "PRAGMA table_info(production_plan)"
            ).fetchall()
        }
        if "production_start_datetime" not in existing_columns:
            self.conn.execute(
                "ALTER TABLE production_plan "
                "ADD COLUMN production_start_datetime TEXT"
            )
        if "production_end_datetime" not in existing_columns:
            self.conn.execute(
                "ALTER TABLE production_plan "
                "ADD COLUMN production_end_datetime TEXT"
            )

        operator_columns = {
            row[1] for row in self.conn.execute(
                "PRAGMA table_info(operator_entries)"
            ).fetchall()
        }
        if "plan_id" not in operator_columns:
            self.conn.execute(
                "ALTER TABLE operator_entries ADD COLUMN plan_id TEXT DEFAULT ''"
            )
        timestamp_columns = {
            "live_entry_date": "TEXT",
            "live_entry_time": "TEXT",
            "entry_timestamp": "TEXT",
            "last_modified_timestamp": "TEXT",
        }
        for column_name, column_type in timestamp_columns.items():
            if column_name not in operator_columns:
                self.conn.execute(
                    f"ALTER TABLE operator_entries "
                    f"ADD COLUMN {column_name} {column_type}"
                )

        # Backfill older entries. created_at may be UTC/server time, so it is
        # retained for traceability while new IST fields are used going forward.
        self.conn.execute(
            """UPDATE operator_entries
               SET entry_timestamp=COALESCE(entry_timestamp, created_at),
                   live_entry_date=COALESCE(
                       live_entry_date, SUBSTR(created_at,1,10)
                   ),
                   live_entry_time=COALESCE(
                       live_entry_time, SUBSTR(created_at,12,8)
                   )
               WHERE entry_timestamp IS NULL
                  OR live_entry_date IS NULL
                  OR live_entry_time IS NULL"""
        )
        self.conn.commit()

    def clear_master_data(self):
        with self.conn:
            for table in (
                "customer_schedules", "stock_demand", "batch_config",
                "process_bom", "machine_recommendations",
                "machine_downtime", "shifts", "breaks",
                "holidays", "weekly_offs", "production_plan"
            ):
                self.conn.execute(f"DELETE FROM {table}")

    def create_database_backup(self, reason="MANUAL"):
        """
        Create a consistent SQLite backup beside the app in Backups/.
        Returns backup metadata.
        """
        backup_dir = self.path.parent / "Backups"
        backup_dir.mkdir(parents=True, exist_ok=True)

        timestamp = india_now()
        safe_reason = re.sub(
            r"[^A-Za-z0-9_-]+", "_", str(reason or "MANUAL")
        ).strip("_") or "MANUAL"
        filename = (
            f"electro_dip_{timestamp:%Y%m%d_%H%M%S}_{safe_reason}.db"
        )
        backup_path = backup_dir / filename

        destination = sqlite3.connect(backup_path)
        try:
            self.conn.backup(destination)
        finally:
            destination.close()

        size_bytes = backup_path.stat().st_size
        with self.conn:
            self.conn.execute(
                """INSERT INTO backup_history
                   (backup_timestamp, backup_reason, backup_filename,
                    backup_size_bytes)
                   VALUES (?, ?, ?, ?)""",
                (
                    timestamp.strftime("%Y-%m-%d %H:%M:%S"),
                    str(reason or "MANUAL"),
                    filename,
                    size_bytes,
                ),
            )
        return {
            "path": str(backup_path),
            "filename": filename,
            "timestamp": timestamp.strftime("%Y-%m-%d %H:%M:%S"),
            "reason": str(reason or "MANUAL"),
            "size_bytes": size_bytes,
        }

    def backup_history_rows(self, limit=50):
        return [
            dict(row)
            for row in self.conn.execute(
                """SELECT * FROM backup_history
                   ORDER BY backup_id DESC LIMIT ?""",
                (int(limit),),
            ).fetchall()
        ]

    def database_health(self):
        backup_row = self.conn.execute(
            """SELECT backup_timestamp, backup_filename
               FROM backup_history
               ORDER BY backup_id DESC LIMIT 1"""
        ).fetchone()
        import_like_rows = self.conn.execute(
            """SELECT COUNT(*) FROM operator_entries
               WHERE source='OPENING_WIP'"""
        ).fetchone()[0]
        return {
            "Database File": self.path.name,
            "Database Size MB": round(
                self.path.stat().st_size / (1024 * 1024), 3
            ) if self.path.exists() else 0,
            "Customer Schedules": self.conn.execute(
                "SELECT COUNT(*) FROM customer_schedules"
            ).fetchone()[0],
            "Production Plan Rows": self.conn.execute(
                "SELECT COUNT(*) FROM production_plan"
            ).fetchone()[0],
            "Operator Entries": self.conn.execute(
                "SELECT COUNT(*) FROM operator_entries"
            ).fetchone()[0],
            "Opening WIP Entries": import_like_rows,
            "Process BOM Rows": self.conn.execute(
                "SELECT COUNT(*) FROM process_bom"
            ).fetchone()[0],
            "Machine Recommendations": self.conn.execute(
                "SELECT COUNT(*) FROM machine_recommendations"
            ).fetchone()[0],
            "Last Backup Time": (
                backup_row["backup_timestamp"] if backup_row else "Never"
            ),
            "Last Backup File": (
                backup_row["backup_filename"] if backup_row else ""
            ),
            "Current IST": india_timestamp_text(),
        }

    def clear_plan(self):
        with self.conn:
            count = self.conn.execute(
                "SELECT COUNT(*) FROM production_plan"
            ).fetchone()[0]
            self.conn.execute("DELETE FROM production_plan")
        return count

    def clear_previous_entries(self):
        with self.conn:
            count = self.conn.execute(
                "SELECT COUNT(*) FROM operator_entries"
            ).fetchone()[0]
            self.conn.execute("DELETE FROM operator_entries")
            self.conn.execute("DELETE FROM production_plan")
        return count

    @staticmethod
    def _iter_rows(ws, max_blank_run=100):
        blank_run = 0
        for row_no, row in enumerate(
            ws.iter_rows(min_row=3, values_only=True), start=3
        ):
            if all(is_blank(v) for v in row):
                blank_run += 1
                if blank_run >= max_blank_run:
                    break
                continue
            blank_run = 0
            yield row_no, row

    def import_workbook(self, source):
        wb = load_workbook(source, data_only=True, read_only=True)
        required = [
            "Customer_Schedules", "Stock_Demand", "Batch_Config",
            "Process_BOM", "Machine_Recommendations"
        ]
        missing = [s for s in required if s not in wb.sheetnames]
        if missing:
            raise ValueError("Missing worksheets: " + ", ".join(missing))

        counts = defaultdict(int)
        warnings = []

        # Master data is replaced, but operator entries are intentionally preserved.
        self.clear_master_data()

        with self.conn:
            ws = wb["Customer_Schedules"]
            for r, row in self._iter_rows(ws):
                try:
                    if any(is_blank(row[i]) for i in (0, 1, 2, 3, 4, 5)):
                        warnings.append(f"Customer_Schedules row {r}: incomplete row skipped")
                        continue
                    due = combine_date_time(row[4], row[5])
                    self.conn.execute(
                        """INSERT OR REPLACE INTO customer_schedules
                        (schedule_id, customer_name, part_name, customer_qty,
                         due_datetime, priority)
                        VALUES (?, ?, ?, ?, ?, ?)""",
                        (
                            str(row[0]).strip(), str(row[1]).strip(),
                            str(row[2]).strip(), float(row[3]),
                            due.isoformat(sep=" "), int(float(row[6] or 999))
                        )
                    )
                    counts["Customer Schedules"] += 1
                except Exception as exc:
                    warnings.append(f"Customer_Schedules row {r}: {exc}")

            ws = wb["Stock_Demand"]
            for r, row in self._iter_rows(ws):
                if is_blank(row[0]):
                    continue
                self.conn.execute(
                    """INSERT OR REPLACE INTO stock_demand
                    (part_name, current_stock, minimum_stock, remarks)
                    VALUES (?, ?, ?, ?)""",
                    (
                        str(row[0]).strip(), float(row[1] or 0),
                        float(row[2] or 0), str(row[3] or "")
                    )
                )
                counts["Stock Records"] += 1

            ws = wb["Batch_Config"]
            for r, row in self._iter_rows(ws):
                if is_blank(row[0]):
                    continue
                self.conn.execute(
                    """INSERT OR REPLACE INTO batch_config
                    (part_name, production_batch, transportation_batch)
                    VALUES (?, ?, ?)""",
                    (
                        str(row[0]).strip(), float(row[1] or 0),
                        float(row[2] or 0)
                    )
                )
                counts["Batch Configurations"] += 1

            ws = wb["Process_BOM"]
            for r, row in self._iter_rows(ws):
                if is_blank(row[0]) or is_blank(row[1]) or is_blank(row[2]):
                    continue
                self.conn.execute(
                    """INSERT INTO process_bom
                    (part_name, process_sequence, operation_name, process_type,
                     cycle_time_sec, setup_time_min, outsource_lead_hours,
                     qty_multiplier, scrap_allowance)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                    (
                        str(row[0]).strip(), int(float(row[1])),
                        str(row[2]).strip(), str(row[3] or "INHOUSE").strip().upper(),
                        float(row[4] or 0), float(row[5] or 0),
                        float(row[6] or 0), max(float(row[7] or 1), 1),
                        float(row[8] or 0)
                    )
                )
                counts["Process BOM Operations"] += 1

            ws = wb["Machine_Recommendations"]
            for r, row in self._iter_rows(ws):
                if is_blank(row[0]) or is_blank(row[1]):
                    continue
                for pref, machine in enumerate(row[2:17], start=1):
                    if is_blank(machine):
                        continue
                    self.conn.execute(
                        """INSERT INTO machine_recommendations
                        (part_name, operation_name, machine_name, preference_order)
                        VALUES (?, ?, ?, ?)""",
                        (
                            str(row[0]).strip(), str(row[1]).strip(),
                            str(machine).strip(), pref
                        )
                    )
                    counts["Machine Recommendations"] += 1

            if "Machine_Downtime" in wb.sheetnames:
                ws = wb["Machine_Downtime"]
                for r, row in self._iter_rows(ws):
                    if is_blank(row[0]):
                        continue
                    if all(is_blank(row[i]) for i in (1, 2, 3, 4)):
                        continue
                    try:
                        start_dt = combine_date_time(row[1], row[2])
                        end_dt = combine_date_time(row[3], row[4])
                        if end_dt <= start_dt:
                            warnings.append(
                                f"Machine_Downtime row {r}: end must be after start"
                            )
                            continue
                        self.conn.execute(
                            """INSERT INTO machine_downtime
                            (machine_name, start_datetime, end_datetime, reason)
                            VALUES (?, ?, ?, ?)""",
                            (
                                str(row[0]).strip(),
                                start_dt.isoformat(sep=" "),
                                end_dt.isoformat(sep=" "),
                                str(row[5] or ""),
                            )
                        )
                        counts["Machine Downtime"] += 1
                    except Exception as exc:
                        warnings.append(f"Machine_Downtime row {r}: {exc}")

            if "Shifts" in wb.sheetnames:
                ws = wb["Shifts"]
                for r, row in self._iter_rows(ws):
                    if is_blank(row[0]) or is_blank(row[1]) or is_blank(row[2]):
                        continue
                    try:
                        self.conn.execute(
                            """INSERT OR REPLACE INTO shifts
                            (shift_name, start_time, end_time, active)
                            VALUES (?, ?, ?, ?)""",
                            (
                                str(row[0]).strip(),
                                parse_time(row[1]).strftime("%H:%M"),
                                parse_time(row[2]).strftime("%H:%M"),
                                1 if str(row[3] or "Y").strip().upper() == "Y" else 0,
                            )
                        )
                        counts["Shifts"] += 1
                    except Exception as exc:
                        warnings.append(f"Shifts row {r}: {exc}")

            if "Breaks" in wb.sheetnames:
                ws = wb["Breaks"]
                for r, row in self._iter_rows(ws):
                    if any(is_blank(row[i]) for i in (0, 1, 2, 3)):
                        continue
                    try:
                        self.conn.execute(
                            """INSERT INTO breaks
                            (shift_name, break_name, start_time, end_time)
                            VALUES (?, ?, ?, ?)""",
                            (
                                str(row[0]).strip(),
                                str(row[1]).strip(),
                                parse_time(row[2]).strftime("%H:%M"),
                                parse_time(row[3]).strftime("%H:%M"),
                            )
                        )
                        counts["Breaks"] += 1
                    except Exception as exc:
                        warnings.append(f"Breaks row {r}: {exc}")

            if "Holidays" in wb.sheetnames:
                ws = wb["Holidays"]
                for r, row in self._iter_rows(ws):
                    if is_blank(row[0]):
                        continue
                    try:
                        self.conn.execute(
                            """INSERT OR REPLACE INTO holidays
                            (holiday_date, holiday_name)
                            VALUES (?, ?)""",
                            (parse_date(row[0]).isoformat(), str(row[1] or ""))
                        )
                        counts["Holidays"] += 1
                    except Exception as exc:
                        warnings.append(f"Holidays row {r}: {exc}")

            if "Weekly_Offs" in wb.sheetnames:
                ws = wb["Weekly_Offs"]
                for r, row in self._iter_rows(ws):
                    if is_blank(row[0]):
                        continue
                    try:
                        day_no = int(float(row[0]))
                        if not 1 <= day_no <= 7:
                            continue
                        self.conn.execute(
                            """INSERT OR REPLACE INTO weekly_offs
                            (day_number, day_name, is_off)
                            VALUES (?, ?, ?)""",
                            (
                                day_no,
                                str(row[1] or ""),
                                1 if str(row[2] or "N").strip().upper() == "Y" else 0,
                            )
                        )
                        counts["Weekly Off Rows"] += 1
                    except Exception as exc:
                        warnings.append(f"Weekly_Offs row {r}: {exc}")

            if "Opening_WIP" in wb.sheetnames:
                ws = wb["Opening_WIP"]
                existing = self.conn.execute(
                    "SELECT COUNT(*) FROM operator_entries"
                ).fetchone()[0]
                # Opening WIP is imported only when no previous entries exist.
                if existing == 0:
                    for r, row in self._iter_rows(ws):
                        if is_blank(row[0]) or is_blank(row[1]) or is_blank(row[2]):
                            continue
                        actual = float(row[4] or 0)
                        rejected = float(row[5] or 0)
                        good = float(row[6] if not is_blank(row[6]) else max(actual-rejected, 0))
                        self.conn.execute(
                            """INSERT INTO operator_entries
                            (entry_date, part_name, process_sequence, operation_name,
                             machine_name, actual_qty, rejected_qty, good_qty,
                             operator_name, remarks, source, created_at)
                            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'OPENING_WIP', ?)""",
                            (
                                parse_date(row[0]).isoformat(),
                                str(row[1]).strip(), int(float(row[2])),
                                str(row[3] or "").strip(), str(row[7] or "").strip(),
                                actual, rejected, good, str(row[8] or "").strip(),
                                str(row[9] or "").strip(),
                                datetime.now().isoformat(sep=" ", timespec="seconds")
                            )
                        )
                        counts["Opening WIP Entries"] += 1
                else:
                    warnings.append(
                        "Opening_WIP was not re-imported because previous operator entries already exist."
                    )

        wb.close()
        return {
            "counts": dict(counts),
            "warnings": warnings,
            "previous_entries_preserved": self.conn.execute(
                "SELECT COUNT(*) FROM operator_entries"
            ).fetchone()[0],
        }

    def personnel_names(self, role_name):
        return [
            row["person_name"]
            for row in self.conn.execute(
                """SELECT person_name
                   FROM personnel_master
                   WHERE role_name=? AND active=1
                   ORDER BY person_name""",
                (str(role_name).strip().upper(),)
            ).fetchall()
        ]

    def add_personnel(self, person_name, role_name):
        name = str(person_name or "").strip()
        role = str(role_name or "").strip().upper()
        if not name:
            raise ValueError("Person name is required.")
        if role not in ("OPERATOR", "SUPERVISOR"):
            raise ValueError("Role must be OPERATOR or SUPERVISOR.")
        with self.conn:
            self.conn.execute(
                """INSERT OR IGNORE INTO personnel_master
                (person_name, role_name, active)
                VALUES (?, ?, 1)""",
                (name, role)
            )

    def plan_id_rows(self):
        return self.production_plan_report_rows()

    def plan_id_detail(self, plan_id):
        row = self.conn.execute(
            """SELECT p.*,
                      COALESCE(s.priority, 999) AS priority,
                      COALESCE(s.customer_qty, 0) AS customer_demand
               FROM production_plan p
               LEFT JOIN customer_schedules s
                 ON p.schedule_id=s.schedule_id
               WHERE p.plan_id=?""",
            (str(plan_id),)
        ).fetchone()
        return dict(row) if row else None

    def machine_dropdown_values(self):
        values = set()
        for row in self.conn.execute(
            "SELECT DISTINCT machine_name FROM machine_recommendations"
        ).fetchall():
            if row["machine_name"]:
                values.add(row["machine_name"])
        for row in self.conn.execute(
            "SELECT DISTINCT machine_name FROM production_plan"
        ).fetchall():
            if row["machine_name"]:
                values.add(row["machine_name"])
        return sorted(values)

    def shift_dropdown_values(self):
        values = [
            row["shift_name"]
            for row in self.conn.execute(
                "SELECT shift_name FROM shifts WHERE active=1 ORDER BY shift_name"
            ).fetchall()
        ]
        return values

    def _plan_route_rows(self, schedule_id, lot_no):
        """Return all process rows for one schedule and transportation lot."""
        return self.conn.execute(
            """SELECT *
               FROM production_plan
               WHERE schedule_id=? AND lot_no=?
               ORDER BY process_sequence""",
            (str(schedule_id), int(lot_no))
        ).fetchall()

    def _operator_actual_sum(self, plan_id):
        row = self.conn.execute(
            """SELECT COALESCE(SUM(actual_qty),0) AS qty
               FROM operator_entries
               WHERE plan_id=?""",
            (str(plan_id),)
        ).fetchone()
        return float(row["qty"] or 0)

    def _operator_good_sum(self, plan_id):
        row = self.conn.execute(
            """SELECT COALESCE(SUM(good_qty),0) AS qty
               FROM operator_entries
               WHERE plan_id=?""",
            (str(plan_id),)
        ).fetchone()
        return float(row["qty"] or 0)

    def sequence_gate_status(self, plan_id):
        """
        Return the maximum quantity that may be entered for a Plan ID.

        First process:
            max allowed = lot planned qty - actual already entered here

        Subsequent process:
            max allowed = previous process good qty - actual already entered here
        """
        plan = self.plan_id_detail(plan_id)
        if not plan:
            raise ValueError("Selected Plan ID was not found.")

        schedule_id = str(plan["schedule_id"])
        lot_no = int(plan["lot_no"])
        current_sequence = int(plan["process_sequence"])
        route = self._plan_route_rows(schedule_id, lot_no)

        current_index = None
        for idx, row in enumerate(route):
            if int(row["process_sequence"]) == current_sequence:
                current_index = idx
                break
        if current_index is None:
            raise ValueError("Current operation is not available in the plan route.")

        already_processed = self._operator_actual_sum(plan_id)
        lot_planned_qty = float(plan["planned_qty"] or 0)

        if current_index == 0:
            previous_plan_id = ""
            previous_operation = "FIRST OPERATION"
            previous_good_qty = lot_planned_qty
            max_allowed = max(lot_planned_qty - already_processed, 0.0)
            gate_open = max_allowed > 0
            reason = (
                "First operation is limited by the lot planned quantity."
                if gate_open else
                "The full planned quantity has already been entered for this operation."
            )
        else:
            previous = route[current_index - 1]
            previous_plan_id = str(previous["plan_id"])
            previous_operation = str(previous["operation_name"])
            previous_good_qty = self._operator_good_sum(previous_plan_id)
            max_allowed = max(previous_good_qty - already_processed, 0.0)
            gate_open = max_allowed > 0

            if previous_good_qty <= 0:
                reason = (
                    f"Operation jumping is blocked. No good quantity is available "
                    f"from previous operation: {previous_operation}."
                )
            elif max_allowed <= 0:
                reason = (
                    f"No balance is available from previous operation: "
                    f"{previous_operation}."
                )
            else:
                reason = (
                    f"Maximum allowed is based on good quantity released by "
                    f"{previous_operation}."
                )

        return {
            "plan_id": str(plan_id),
            "schedule_id": schedule_id,
            "lot_no": lot_no,
            "part_name": plan["part_name"],
            "current_sequence": current_sequence,
            "current_operation": plan["operation_name"],
            "lot_planned_qty": lot_planned_qty,
            "previous_plan_id": previous_plan_id,
            "previous_operation": previous_operation,
            "previous_good_qty": previous_good_qty,
            "already_processed_here": already_processed,
            "maximum_entry_allowed": max_allowed,
            "gate_open": gate_open,
            "reason": reason,
        }

    def validate_sequence_entry(self, plan_id, actual_qty):
        status = self.sequence_gate_status(plan_id)
        actual = max(float(actual_qty or 0), 0.0)

        if actual <= 0:
            raise ValueError("Actual Qty must be greater than zero.")

        if not status["gate_open"]:
            raise ValueError(status["reason"])

        if actual > status["maximum_entry_allowed"] + 1e-9:
            raise ValueError(
                f"Cannot save {actual:g} pcs. Maximum allowed for this operation "
                f"and lot is {status['maximum_entry_allowed']:g} pcs. "
                f"Previous operation good qty: {status['previous_good_qty']:g}; "
                f"already processed here: {status['already_processed_here']:g}."
            )

        return status

    def add_operator_entry(
        self, entry_date, part_name, process_sequence, operation_name,
        actual_qty, rejected_qty, machine_name="", shift_name="",
        plan_id="", schedule_id="", customer_name="", planned_qty=0, status="",
        operator_name="", supervisor_name="", remarks=""
    ):
        actual = max(float(actual_qty or 0), 0)
        rejected = max(float(rejected_qty or 0), 0)

        if rejected > actual:
            raise ValueError("Rejected Qty cannot exceed Actual Qty.")

        # Sequence gate applies to production-plan entries. Opening WIP and
        # administrative records without a Plan ID are excluded.
        if str(plan_id or "").strip():
            self.validate_sequence_entry(plan_id, actual)

        good = max(actual - rejected, 0)
        captured_ist = india_now()
        captured_timestamp = captured_ist.strftime("%Y-%m-%d %H:%M:%S")
        with self.conn:
            self.conn.execute(
                """INSERT INTO operator_entries
                (plan_id, entry_date, schedule_id, customer_name, part_name,
                 process_sequence, operation_name, machine_name, shift_name,
                 planned_qty, actual_qty, rejected_qty, good_qty, status,
                 operator_name, supervisor_name, remarks, source,
                 live_entry_date, live_entry_time, entry_timestamp,
                 last_modified_timestamp, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?,
                        'MANUAL', ?, ?, ?, NULL, ?)""",
                (
                    str(plan_id or ""),
                    parse_date(entry_date).isoformat(), str(schedule_id or ""),
                    str(customer_name or ""), str(part_name).strip(),
                    int(process_sequence), str(operation_name).strip(),
                    str(machine_name or ""), str(shift_name or ""),
                    float(planned_qty or 0), actual, rejected, good,
                    str(status or ""), str(operator_name or ""),
                    str(supervisor_name or ""), str(remarks or ""),
                    captured_ist.strftime("%Y-%m-%d"),
                    captured_ist.strftime("%H:%M:%S"),
                    captured_timestamp,
                    captured_timestamp,
                )
            )
        return good

    def operator_entry_by_id(self, entry_id):
        row = self.conn.execute(
            "SELECT * FROM operator_entries WHERE entry_id=?",
            (int(entry_id),),
        ).fetchone()
        return dict(row) if row else None

    def update_operator_entry(
        self,
        entry_id,
        actual_qty,
        rejected_qty,
        status="",
        operator_name="",
        supervisor_name="",
        remarks="",
    ):
        """
        Edit an existing operator entry.

        Plan-controlled fields remain unchanged:
        Plan ID, part, process, machine, shift, planned qty and entry date.
        """
        existing = self.operator_entry_by_id(entry_id)
        if not existing:
            raise ValueError("Operator entry was not found.")

        actual = max(float(actual_qty or 0), 0.0)
        rejected = max(float(rejected_qty or 0), 0.0)

        if actual <= 0:
            raise ValueError("Actual Qty must be greater than zero.")
        if rejected > actual:
            raise ValueError("Rejected Qty cannot exceed Actual Qty.")

        # Recalculate sequence gate while excluding the current entry,
        # otherwise the existing quantity would be counted twice.
        plan_id = str(existing.get("plan_id") or "")
        if plan_id:
            original_actual = float(existing.get("actual_qty") or 0)
            gate = self.sequence_gate_status(plan_id)
            maximum_if_editing = (
                float(gate["maximum_entry_allowed"]) + original_actual
            )
            if actual > maximum_if_editing + 1e-9:
                raise ValueError(
                    f"Cannot save {actual:g} pcs. Maximum allowed after "
                    f"considering the previous operation is "
                    f"{maximum_if_editing:g} pcs."
                )

        good = max(actual - rejected, 0.0)

        with self.conn:
            self.conn.execute(
                """UPDATE operator_entries
                   SET actual_qty=?,
                       rejected_qty=?,
                       good_qty=?,
                       status=?,
                       operator_name=?,
                       supervisor_name=?,
                       remarks=?,
                       last_modified_timestamp=?
                   WHERE entry_id=?""",
                (
                    actual,
                    rejected,
                    good,
                    str(status or ""),
                    str(operator_name or ""),
                    str(supervisor_name or ""),
                    str(remarks or ""),
                    india_timestamp_text(),
                    int(entry_id),
                ),
            )
        return good

    def delete_operator_entry(self, entry_id):
        """
        Permanently delete one operator entry.

        All WIP, ageing and sequence-gate reports recalculate automatically
        because they are derived from operator_entries.
        """
        existing = self.operator_entry_by_id(entry_id)
        if not existing:
            raise ValueError("Operator entry was not found.")

        with self.conn:
            self.conn.execute(
                "DELETE FROM operator_entries WHERE entry_id=?",
                (int(entry_id),),
            )
        return existing

    def operator_entries(self):
        return self.conn.execute(
            """SELECT * FROM operator_entries
               ORDER BY COALESCE(entry_timestamp, created_at) DESC,
                        entry_id DESC"""
        ).fetchall()

    def bom_rows(self):
        return self.conn.execute(
            """SELECT * FROM process_bom
               ORDER BY part_name, process_sequence"""
        ).fetchall()

    def part_final_sequence(self, part_name):
        row = self.conn.execute(
            """SELECT MAX(process_sequence) AS seq
               FROM process_bom WHERE part_name=?""",
            (part_name,)
        ).fetchone()
        return int(row["seq"]) if row and row["seq"] is not None else None

    def process_cumulative_good(self, part_name):
        rows = self.conn.execute(
            """SELECT process_sequence, SUM(good_qty) AS good
               FROM operator_entries
               WHERE part_name=?
               GROUP BY process_sequence""",
            (part_name,)
        ).fetchall()
        return {int(r["process_sequence"]): float(r["good"] or 0) for r in rows}

    def wip_rows(self):
        """
        Return physical WIP by process.

        Daily entries are cumulative good quantities. To remain physically
        consistent, an upstream process can never have less cumulative output
        than a downstream process. Missing intermediate entries are therefore
        normalized from downstream to upstream before WIP is calculated.
        """
        output = []
        parts = self.conn.execute(
            "SELECT DISTINCT part_name FROM process_bom ORDER BY part_name"
        ).fetchall()
        for part_row in parts:
            part = part_row["part_name"]
            bom = self.conn.execute(
                """SELECT process_sequence, operation_name
                   FROM process_bom WHERE part_name=?
                   ORDER BY process_sequence""",
                (part,)
            ).fetchall()
            raw_good = self.process_cumulative_good(part)

            effective = {}
            downstream_floor = 0.0
            for row in reversed(bom):
                seq = int(row["process_sequence"])
                downstream_floor = max(raw_good.get(seq, 0.0), downstream_floor)
                effective[seq] = downstream_floor

            for idx, row in enumerate(bom):
                seq = int(row["process_sequence"])
                cumulative = effective.get(seq, 0.0)
                if idx + 1 < len(bom):
                    next_seq = int(bom[idx+1]["process_sequence"])
                    next_cumulative = effective.get(next_seq, 0.0)
                    wip = max(cumulative - next_cumulative, 0.0)
                else:
                    wip = 0.0

                output.append({
                    "part_name": part,
                    "process_sequence": seq,
                    "operation_name": row["operation_name"],
                    "reported_cumulative_good_qty": raw_good.get(seq, 0.0),
                    "normalized_cumulative_good_qty": cumulative,
                    "wip_after_process": wip,
                    "is_dispatch_sequence": idx == len(bom)-1,
                })
        return output

    def total_dispatched_by_part(self, part_name):
        final_seq = self.part_final_sequence(part_name)
        if final_seq is None:
            return 0.0
        row = self.conn.execute(
            """SELECT COALESCE(SUM(good_qty),0) AS qty
               FROM operator_entries
               WHERE part_name=? AND process_sequence=?""",
            (part_name, final_seq)
        ).fetchone()
        return float(row["qty"] or 0)

    def schedule_balance_rows(self):
        rows = []
        parts = self.conn.execute(
            """SELECT DISTINCT part_name
               FROM customer_schedules ORDER BY part_name"""
        ).fetchall()

        for part_row in parts:
            part = part_row["part_name"]
            schedules = self.conn.execute(
                """SELECT * FROM customer_schedules
                   WHERE part_name=?
                   ORDER BY due_datetime, priority, schedule_id""",
                (part,)
            ).fetchall()
            stock = self.conn.execute(
                "SELECT * FROM stock_demand WHERE part_name=?",
                (part,)
            ).fetchone()
            current_stock = float(stock["current_stock"] or 0) if stock else 0.0
            minimum_stock = float(stock["minimum_stock"] or 0) if stock else 0.0
            dispatched_pool = self.total_dispatched_by_part(part)
            stock_pool = current_stock

            for sch in schedules:
                demand = float(sch["customer_qty"] or 0)
                stock_alloc = min(stock_pool, demand)
                stock_pool -= stock_alloc
                after_stock = demand - stock_alloc
                dispatch_alloc = min(dispatched_pool, after_stock)
                dispatched_pool -= dispatch_alloc
                remaining = max(after_stock - dispatch_alloc, 0)
                rows.append({
                    "requirement_type": "CUSTOMER",
                    "schedule_id": sch["schedule_id"],
                    "customer_name": sch["customer_name"],
                    "part_name": part,
                    "due_datetime": sch["due_datetime"],
                    "customer_demand": demand,
                    "allocated_current_stock": stock_alloc,
                    "allocated_dispatch_qty": dispatch_alloc,
                    "minimum_stock_qty": 0.0,
                    "revised_requirement": remaining,
                    "priority": sch["priority"],
                })

            # Minimum stock is added once, only after all customer schedules.
            remaining_stock_for_min = stock_pool
            remaining_dispatch_for_min = dispatched_pool
            min_after_stock = max(minimum_stock - remaining_stock_for_min, 0)
            min_after_dispatch = max(min_after_stock - remaining_dispatch_for_min, 0)
            rows.append({
                "requirement_type": "MINIMUM_STOCK",
                "schedule_id": f"MIN-STOCK-{part}",
                "customer_name": "",
                "part_name": part,
                "due_datetime": None,
                "customer_demand": 0.0,
                "allocated_current_stock": min(remaining_stock_for_min, minimum_stock),
                "allocated_dispatch_qty": min(remaining_dispatch_for_min, min_after_stock),
                "minimum_stock_qty": minimum_stock,
                "revised_requirement": min_after_dispatch,
                "priority": 9999,
            })
        return rows

    def downstream_available_qty(self, part_name, process_sequence):
        """
        Physical WIP already completed at this process or downstream,
        excluding dispatched quantity because dispatch is already deducted
        from customer schedules.
        """
        return sum(
            float(row["wip_after_process"] or 0)
            for row in self.wip_rows()
            if row["part_name"] == part_name
            and int(row["process_sequence"]) >= int(process_sequence)
            and not row["is_dispatch_sequence"]
        )


    def opening_wip_total(self, part_name):
        return sum(
            float(row["wip_after_process"] or 0)
            for row in self.wip_rows()
            if row["part_name"] == part_name
            and not row["is_dispatch_sequence"]
        )

    def schedule_line_calculation_rows(self):
        """
        Schedule calculation without WIP deduction.

        Schedule Plan Qty =
            Customer Demand + Minimum Stock Level - Allocated Current FG Stock

        Important:
        - Opening WIP is not deducted.
        - Previous-schedule carry-forward WIP is not deducted.
        - Process WIP is handled only after schedule quantity calculation.
        - Current finished-goods stock is allocated FIFO because it is FG stock,
          not process WIP.
        """
        results = []
        part_rows = self.conn.execute(
            """SELECT DISTINCT part_name
               FROM customer_schedules
               ORDER BY part_name"""
        ).fetchall()

        for part_row in part_rows:
            part = part_row["part_name"]
            schedules = self.conn.execute(
                """SELECT * FROM customer_schedules
                   WHERE part_name=?
                   ORDER BY due_datetime, priority, schedule_id""",
                (part,)
            ).fetchall()

            stock = self.conn.execute(
                "SELECT * FROM stock_demand WHERE part_name=?",
                (part,)
            ).fetchone()

            minimum_stock = float(stock["minimum_stock"] or 0) if stock else 0.0
            current_stock_pool = float(stock["current_stock"] or 0) if stock else 0.0

            for line_no, schedule in enumerate(schedules, start=1):
                demand = float(schedule["customer_qty"] or 0)
                gross_requirement = demand + minimum_stock

                allocated_fg_stock = min(current_stock_pool, gross_requirement)
                current_stock_pool = max(current_stock_pool - allocated_fg_stock, 0.0)

                plan_qty = max(gross_requirement - allocated_fg_stock, 0.0)

                results.append({
                    "line_no": line_no,
                    "schedule_id": schedule["schedule_id"],
                    "customer_name": schedule["customer_name"],
                    "part_name": part,
                    "due_datetime": schedule["due_datetime"],
                    "priority": schedule["priority"],
                    "customer_demand": demand,
                    "minimum_stock_level": minimum_stock,
                    "gross_requirement": gross_requirement,
                    "allocated_current_fg_stock": allocated_fg_stock,
                    "opening_wip_stock": 0.0,
                    "previous_schedule_carry_forward_wip": 0.0,
                    "wip_used_for_line": 0.0,
                    "plan_qty": plan_qty,
                    "carry_forward_wip_after_line": 0.0,
                    "calculation_rule": (
                        "Demand + Minimum Stock - Allocated Current FG Stock; "
                        "WIP not considered in schedule calculation"
                    ),
                })

        return results


    def process_schedule_wip_rows(self):
        output = []
        schedule_rows = self.schedule_line_calculation_rows()
        by_part = defaultdict(list)
        for row in schedule_rows:
            by_part[row["part_name"]].append(row)

        for part, schedules in by_part.items():
            bom = self.conn.execute(
                """SELECT * FROM process_bom WHERE part_name=?
                   ORDER BY process_sequence""", (part,)
            ).fetchall()
            pools = {
                int(op["process_sequence"]): self.downstream_available_qty(
                    part, int(op["process_sequence"])
                ) for op in bom
            }
            for schedule in schedules:
                line_required = float(schedule["plan_qty"] or 0)
                for op in bom:
                    seq = int(op["process_sequence"])
                    available = pools.get(seq, 0.0)
                    wip_used = min(available, line_required)
                    net_process_plan = max(line_required - wip_used, 0.0)
                    pools[seq] = max(available - line_required, 0.0)
                    output.append({
                        "line_no": schedule["line_no"],
                        "schedule_id": schedule["schedule_id"],
                        "customer_name": schedule["customer_name"],
                        "part_name": part,
                        "due_datetime": schedule["due_datetime"],
                        "process_sequence": seq,
                        "operation_name": op["operation_name"],
                        "schedule_plan_qty": line_required,
                        "process_wip_available_before": available,
                        "process_wip_used": wip_used,
                        "net_process_plan_qty": net_process_plan,
                        "process_wip_carry_forward": pools[seq],
                    })
        return output

    @staticmethod
    def split_transportation_lots(total_qty, transportation_batch):
        """
        Split the exact calculated quantity into transportation lots.
        The total quantity is never rounded or changed.

        Example:
            240 qty, batch 100 -> [100, 100, 40]
        """
        total = max(float(total_qty or 0), 0.0)
        batch = float(transportation_batch or 0)

        if total <= 0:
            return []
        if batch <= 0:
            return [total]

        lots = []
        remaining = total
        while remaining > 1e-9:
            lot_qty = min(batch, remaining)
            lots.append(lot_qty)
            remaining -= lot_qty
        return lots

    def generate_plan(self):
        """
        Build a WIP-adjusted, transportation-batch plan and backward-schedule
        every in-house operation using cycle/setup times and factory calendar.
        """
        schedule_rows = self.schedule_line_calculation_rows()
        process_map = {
            (row["schedule_id"], int(row["process_sequence"])): row
            for row in self.process_schedule_wip_rows()
        }

        shifts = [
            (
                row["shift_name"],
                parse_time(row["start_time"]),
                parse_time(row["end_time"]),
            )
            for row in self.conn.execute(
                "SELECT * FROM shifts WHERE active=1"
            ).fetchall()
        ]
        if not shifts:
            shifts = [("Shift-A", time(7, 0), time(19, 0))]

        break_map = defaultdict(list)
        for row in self.conn.execute("SELECT * FROM breaks").fetchall():
            break_map[row["shift_name"]].append(
                (parse_time(row["start_time"]), parse_time(row["end_time"]))
            )

        holidays = {
            parse_date(row["holiday_date"])
            for row in self.conn.execute("SELECT * FROM holidays").fetchall()
        }
        weekly_offs = {
            int(row["day_number"]) - 1
            for row in self.conn.execute(
                "SELECT * FROM weekly_offs WHERE is_off=1"
            ).fetchall()
        }

        downtime = defaultdict(list)
        for row in self.conn.execute("SELECT * FROM machine_downtime").fetchall():
            downtime[row["machine_name"]].append(
                (
                    datetime.fromisoformat(row["start_datetime"]),
                    datetime.fromisoformat(row["end_datetime"]),
                )
            )

        reserved = defaultdict(set)

        def in_clock_interval(current_time, start_time, end_time):
            if start_time < end_time:
                return start_time <= current_time < end_time
            return current_time >= start_time or current_time < end_time

        def shift_for(moment):
            current = moment.time()
            for shift_name, start_time, end_time in shifts:
                if in_clock_interval(current, start_time, end_time):
                    return shift_name
            return ""

        def working_minute(moment):
            if moment.date() in holidays or moment.weekday() in weekly_offs:
                return False
            shift_name = shift_for(moment)
            if not shift_name:
                return False
            for break_start, break_end in break_map.get(shift_name, []):
                if in_clock_interval(moment.time(), break_start, break_end):
                    return False
            return True

        def machine_available(machine, moment):
            return not any(
                start <= moment < end
                for start, end in downtime.get(machine, [])
            )

        def find_backward_slot(machine, finish_by, required_minutes):
            """
            Find the latest available minute block before finish_by.
            Non-working minutes may separate working minutes; the returned
            start/end represent the calendar span containing required work.
            """
            cursor = finish_by.replace(second=0, microsecond=0) - timedelta(minutes=1)
            limit = cursor - timedelta(days=730)
            selected = []

            while cursor >= limit and len(selected) < required_minutes:
                if (
                    working_minute(cursor)
                    and machine_available(machine, cursor)
                    and cursor not in reserved[machine]
                ):
                    selected.append(cursor)
                cursor -= timedelta(minutes=1)

            if len(selected) < required_minutes:
                return None

            selected.sort()
            start_dt = selected[0]
            end_dt = selected[-1] + timedelta(minutes=1)
            return start_dt, end_dt, selected, shift_for(start_dt)

        with self.conn:
            self.conn.execute("DELETE FROM production_plan")

            # Earliest due schedules receive capacity first.
            for schedule in schedule_rows:
                if float(schedule["plan_qty"] or 0) <= 0:
                    continue

                part = schedule["part_name"]
                batch = self.conn.execute(
                    "SELECT * FROM batch_config WHERE part_name=?",
                    (part,)
                ).fetchone()
                transportation_batch = (
                    float(batch["transportation_batch"] or 0)
                    if batch else 0.0
                )
                if transportation_batch <= 0:
                    transportation_batch = float(schedule["plan_qty"])

                bom = self.conn.execute(
                    """SELECT * FROM process_bom
                       WHERE part_name=?
                       ORDER BY process_sequence DESC""",
                    (part,)
                ).fetchall()

                # Process-specific required quantities and lots.
                operation_lots = {}
                max_lots = 0
                for op in bom:
                    seq = int(op["process_sequence"])
                    process_row = process_map.get((schedule["schedule_id"], seq))
                    process_required = (
                        float(process_row["net_process_plan_qty"] or 0)
                        if process_row
                        else float(schedule["plan_qty"])
                    )
                    lots = self.split_transportation_lots(
                        process_required,
                        transportation_batch,
                    )
                    operation_lots[seq] = lots
                    max_lots = max(max_lots, len(lots))

                due_dt = datetime.fromisoformat(schedule["due_datetime"])

                # Schedule each transportation lot backward through the route.
                for lot_index in range(max_lots):
                    next_operation_start = due_dt

                    for op in bom:
                        seq = int(op["process_sequence"])
                        lots = operation_lots.get(seq, [])
                        if lot_index >= len(lots):
                            continue

                        lot_qty = float(lots[lot_index])
                        process_row = process_map.get((schedule["schedule_id"], seq))
                        process_wip_used = (
                            float(process_row["process_wip_used"] or 0)
                            if process_row else 0.0
                        )
                        process_qty = (
                            lot_qty
                            * max(float(op["qty_multiplier"] or 1), 1)
                            * (1 + float(op["scrap_allowance"] or 0) / 100)
                        )

                        process_type = str(op["process_type"] or "INHOUSE").upper()
                        note = (
                            f"Gross {schedule['gross_requirement']:g}; "
                            "schedule WIP not considered; "
                            f"process WIP used {process_wip_used:g}; "
                            "backward scheduled"
                        )

                        if process_type == "OUTSOURCE":
                            end_dt = next_operation_start
                            start_dt = end_dt - timedelta(
                                hours=float(op["outsource_lead_hours"] or 0)
                            )
                            machine = "OUTSOURCE-" + str(op["operation_name"])
                            shift_name = "OUTSOURCE"
                        else:
                            required_minutes = max(
                                1,
                                math.ceil(
                                    process_qty * float(op["cycle_time_sec"] or 0) / 60
                                    + float(op["setup_time_min"] or 0)
                                )
                            )

                            machines = self.conn.execute(
                                """SELECT machine_name, preference_order
                                   FROM machine_recommendations
                                   WHERE part_name=? AND operation_name=?
                                   ORDER BY preference_order""",
                                (part, op["operation_name"])
                            ).fetchall()

                            best = None
                            for machine_row in machines:
                                candidate = find_backward_slot(
                                    machine_row["machine_name"],
                                    next_operation_start,
                                    required_minutes,
                                )
                                if candidate is None:
                                    continue
                                start_candidate, end_candidate, minutes, shift_candidate = candidate
                                score = (
                                    start_candidate,
                                    -int(machine_row["preference_order"]),
                                )
                                if best is None or score > best[0]:
                                    best = (
                                        score,
                                        machine_row["machine_name"],
                                        start_candidate,
                                        end_candidate,
                                        minutes,
                                        shift_candidate,
                                    )

                            if best is None:
                                machine = ""
                                start_dt = None
                                end_dt = None
                                shift_name = ""
                                note += "; no feasible machine capacity"
                            else:
                                _, machine, start_dt, end_dt, minutes, shift_name = best
                                reserved[machine].update(minutes)

                        plan_id = (
                            f"{schedule['schedule_id']}-"
                            f"S{seq:03d}-L{lot_index+1:03d}"
                        )

                        self.conn.execute(
                            """INSERT INTO production_plan
                            (plan_id, requirement_type, schedule_id,
                             customer_name, part_name, process_sequence,
                             operation_name, machine_name, shift_name,
                             planned_qty, lot_no, due_datetime,
                             production_start_datetime,
                             production_end_datetime, note)
                            VALUES (?, 'CUSTOMER_AND_MIN_STOCK', ?, ?, ?, ?,
                                    ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                            (
                                plan_id,
                                schedule["schedule_id"],
                                schedule["customer_name"],
                                part,
                                seq,
                                op["operation_name"],
                                machine,
                                shift_name,
                                lot_qty,
                                lot_index + 1,
                                schedule["due_datetime"],
                                start_dt.isoformat(sep=" ") if start_dt else None,
                                end_dt.isoformat(sep=" ") if end_dt else None,
                                note,
                            )
                        )

                        if start_dt is not None:
                            next_operation_start = start_dt

        return self.conn.execute(
            "SELECT COUNT(*) FROM production_plan"
        ).fetchone()[0]

    def plan_rows(self):
        return self.conn.execute(
            """SELECT * FROM production_plan
               ORDER BY COALESCE(production_start_datetime, due_datetime, '9999-12-31'),
                        machine_name, part_name, process_sequence, lot_no"""
        ).fetchall()


    def process_wip_report_rows(self):
        results = []
        summaries = {
            (r["part_name"], int(r["process_sequence"])): r
            for r in self.wip_rows()
        }
        for bom in self.conn.execute(
            """SELECT part_name, process_sequence, operation_name
               FROM process_bom
               ORDER BY part_name, process_sequence"""
        ).fetchall():
            part = bom["part_name"]
            seq = int(bom["process_sequence"])
            summary = summaries.get((part, seq), {})
            wip_qty = float(summary.get("wip_after_process", 0) or 0)

            oldest = self.conn.execute(
                """SELECT * FROM operator_entries
                   WHERE part_name=? AND process_sequence=?
                   ORDER BY entry_date, entry_id LIMIT 1""",
                (part, seq)
            ).fetchone()
            latest = self.conn.execute(
                """SELECT * FROM operator_entries
                   WHERE part_name=? AND process_sequence=?
                   ORDER BY entry_date DESC, entry_id DESC LIMIT 1""",
                (part, seq)
            ).fetchone()

            oldest_date = oldest["entry_date"] if oldest else None
            age_days = (
                max((date.today() - parse_date(oldest_date)).days, 0)
                if oldest_date and wip_qty > 0 else 0
            )
            age_status = (
                "No WIP" if wip_qty <= 0 else
                "Fresh" if age_days <= 2 else
                "Monitor" if age_days <= 7 else
                "Old WIP"
            )

            customer = latest["customer_name"] if latest and latest["customer_name"] else ""
            if not customer:
                cust = self.conn.execute(
                    """SELECT customer_name FROM customer_schedules
                       WHERE part_name=?
                       ORDER BY due_datetime, priority LIMIT 1""",
                    (part,)
                ).fetchone()
                customer = cust["customer_name"] if cust else ""

            results.append({
                "customer_name": customer,
                "part_name": part,
                "process_sequence": seq,
                "operation_name": bom["operation_name"],
                "reported_good_qty": float(summary.get("reported_cumulative_good_qty", 0) or 0),
                "normalized_good_qty": float(summary.get("normalized_cumulative_good_qty", 0) or 0),
                "wip_qty": wip_qty,
                "machine_name": latest["machine_name"] if latest else "",
                "operator_name": latest["operator_name"] if latest else "",
                "oldest_entry_date": oldest_date,
                "last_updated_date": latest["entry_date"] if latest else None,
                "wip_age_days": age_days,
                "age_status": age_status,
                "is_dispatch_sequence": bool(summary.get("is_dispatch_sequence", False)),
            })
        return results

    def machine_wip_rows(self):
        grouped = defaultdict(float)
        for row in self.process_wip_report_rows():
            if row["wip_qty"] > 0:
                grouped[(row["machine_name"] or "UNASSIGNED", row["operation_name"])] += row["wip_qty"]
        return [
            {"machine_name": k[0], "operation_name": k[1], "wip_qty": v}
            for k, v in sorted(grouped.items())
        ]

    def part_wip_rows(self):
        grouped = defaultdict(float)
        for row in self.process_wip_report_rows():
            grouped[row["part_name"]] += row["wip_qty"]
        return [{"part_name": k, "total_wip_qty": v} for k, v in sorted(grouped.items())]

    def customer_wip_rows(self):
        grouped = defaultdict(float)
        for row in self.process_wip_report_rows():
            grouped[row["customer_name"] or "UNALLOCATED"] += row["wip_qty"]
        return [{"customer_name": k, "total_wip_qty": v} for k, v in sorted(grouped.items())]

    def dispatch_history_rows(self):
        return [
            dict(row) for row in self.conn.execute(
                """SELECT e.* FROM operator_entries e
                   WHERE e.process_sequence=(
                       SELECT MAX(p.process_sequence)
                       FROM process_bom p WHERE p.part_name=e.part_name
                   )
                   ORDER BY e.entry_date DESC, e.entry_id DESC"""
            ).fetchall()
        ]


    def production_plan_report_rows(self):
        """
        Production plan rows with V14 schedule qty and transportation-lot details.
        """
        schedule_qty = {
            row["schedule_id"]: float(row["plan_qty"] or 0)
            for row in self.schedule_line_calculation_rows()
        }
        process_qty = {
            (row["schedule_id"], int(row["process_sequence"])): float(
                row["net_process_plan_qty"] or 0
            )
            for row in self.process_schedule_wip_rows()
        }

        rows = self.conn.execute(
            """SELECT p.*,
                      COALESCE(s.priority, 999) AS priority,
                      COALESCE(s.customer_qty, 0) AS customer_demand,
                      COALESCE(b.transportation_batch, 0) AS transportation_batch
               FROM production_plan p
               LEFT JOIN customer_schedules s
                 ON p.schedule_id=s.schedule_id
               LEFT JOIN batch_config b
                 ON p.part_name=b.part_name
               ORDER BY COALESCE(
                            p.production_start_datetime,
                            p.due_datetime,
                            '9999-12-31'
                        ),
                        p.machine_name, p.part_name,
                        p.process_sequence, p.lot_no"""
        ).fetchall()

        output = []
        for row in rows:
            record = dict(row)
            schedule_id = record["schedule_id"]
            sequence = int(record["process_sequence"])
            record["schedule_plan_qty_v14"] = schedule_qty.get(schedule_id, 0.0)
            record["net_process_plan_qty"] = process_qty.get(
                (schedule_id, sequence),
                record["schedule_plan_qty_v14"],
            )
            record["transportation_batch_qty"] = float(
                record.pop("transportation_batch", 0) or 0
            )
            record["lot_planned_qty"] = float(record.get("planned_qty") or 0)
            output.append(record)
        return output

    def export_production_plan_excel(self, output_path):
        wb = xlsxwriter.Workbook(output_path)
        title = wb.add_format({
            "bold": True, "font_color": "white", "bg_color": "#1F4E78",
            "font_size": 14, "align": "center", "valign": "vcenter"
        })
        header = wb.add_format({
            "bold": True, "font_color": "white", "bg_color": "#5B9BD5",
            "border": 1, "align": "center", "valign": "vcenter",
            "text_wrap": True
        })
        cell = wb.add_format({"border": 1})
        qty = wb.add_format({"border": 1, "num_format": "0.00"})

        rows = self.production_plan_report_rows()
        ws = wb.add_worksheet("Production_Plan")
        if not rows:
            ws.write("A1", "No production plan generated.")
            wb.close()
            return

        columns = [
            "plan_id", "requirement_type", "schedule_id", "customer_name",
            "part_name", "process_sequence", "operation_name", "machine_name",
            "schedule_plan_qty_v14", "net_process_plan_qty",
            "transportation_batch_qty", "lot_no", "lot_planned_qty",
            "production_start_datetime", "production_end_datetime",
            "shift_name", "due_datetime", "priority",
            "customer_demand", "note"
        ]

        ws.merge_range(0, 0, 0, len(columns)-1,
                       "ELECTRO-DIP PRODUCTION PLAN", title)
        for c, col in enumerate(columns):
            ws.write(1, c, col, header)

        for r, row in enumerate(rows, start=2):
            for c, col in enumerate(columns):
                value = row.get(col)
                ws.write(r, c, value, qty if isinstance(value, float) else cell)

        ws.autofilter(1, 0, len(rows)+1, len(columns)-1)
        ws.freeze_panes(2, 0)
        widths = {
            "plan_id": 28, "requirement_type": 22, "schedule_id": 18,
            "customer_name": 22, "part_name": 18, "process_sequence": 16,
            "operation_name": 28, "machine_name": 20,
            "schedule_plan_qty_v14": 20,
            "net_process_plan_qty": 20,
            "transportation_batch_qty": 22,
            "lot_no": 10, "lot_planned_qty": 17,
            "due_datetime": 22,
            "production_start_datetime": 22,
            "production_end_datetime": 22,
            "shift_name": 14, "priority": 10,
            "customer_demand": 18, "note": 45
        }
        for c, col in enumerate(columns):
            ws.set_column(c, c, widths.get(col, 16))
        wb.close()

    def operator_slip_rows(self, due_date=None, machine_name=None):
        rows = self.production_plan_report_rows()
        output = []
        for row in rows:
            if due_date:
                comparison_value = (
                    row.get("production_start_datetime")
                    or row.get("due_datetime")
                )
                if not comparison_value:
                    continue
                if parse_date(comparison_value) != parse_date(due_date):
                    continue
            if machine_name and row.get("machine_name") != machine_name:
                continue
            output.append(row)
        return output

    def operator_slip_dates(self):
        dates = []
        for row in self.production_plan_report_rows():
            comparison_value = (
                row.get("production_start_datetime")
                or row.get("due_datetime")
            )
            if comparison_value:
                d = parse_date(comparison_value)
                if d not in dates:
                    dates.append(d)
        return sorted(dates)

    def operator_slip_machines(self, due_date=None):
        machines = set()
        for row in self.operator_slip_rows(due_date=due_date):
            if row.get("machine_name"):
                machines.add(row["machine_name"])
        return sorted(machines)

    def create_operator_slips_pdf(self, output_path, due_date=None, machine_name=None):
        """
        Create one approved A4 landscape operator slip per machine and due date.
        Multiple planned operations are printed on the same slip.
        """
        rows = self.operator_slip_rows(
            due_date=due_date,
            machine_name=machine_name,
        )
        if not rows:
            raise ValueError("No production-plan rows available for the selected slip filter.")

        groups = defaultdict(list)
        for row in rows:
            slip_date = (
                parse_date(row["production_start_datetime"])
                if row.get("production_start_datetime")
                else (
                    parse_date(row["due_datetime"])
                    if row.get("due_datetime")
                    else date.today()
                )
            )
            machine = row.get("machine_name") or "UNASSIGNED"
            groups[(slip_date, machine)].append(row)

        page_w, page_h = landscape(A4)
        c = canvas.Canvas(str(output_path), pagesize=(page_w, page_h))
        c.setTitle("ELECTRO-DIP Operator Machine Slips")
        logo_path = Path(__file__).resolve().parent / "electro_dip_logo.png"

        for slip_index, ((slip_date, machine), group_rows) in enumerate(
            sorted(groups.items(), key=lambda item: (item[0][0], item[0][1])),
            start=1,
        ):
            group_rows.sort(
                key=lambda r: (
                    int(r.get("process_sequence") or 0),
                    int(r.get("lot_no") or 0),
                    str(r.get("part_name") or ""),
                )
            )

            c.setFillColor(colors.HexColor("#F7F9FB"))
            c.rect(0, 0, page_w, page_h, stroke=0, fill=1)
            c.setStrokeColor(colors.HexColor("#1F4E78"))
            c.setLineWidth(1.2)
            c.roundRect(
                5*mm, 5*mm, page_w-10*mm, page_h-10*mm,
                3*mm, stroke=1, fill=0
            )

            # Header.
            c.setFillColor(colors.HexColor("#1F4E78"))
            c.roundRect(
                5*mm, page_h-30*mm, page_w-10*mm, 25*mm,
                3*mm, stroke=0, fill=1
            )
            if logo_path.exists():
                c.setFillColor(colors.white)
                c.rect(9*mm, page_h-26*mm, 18*mm, 17*mm, stroke=0, fill=1)
                c.drawImage(
                    str(logo_path), 10*mm, page_h-25*mm,
                    width=16*mm, height=15*mm,
                    preserveAspectRatio=True, mask="auto"
                )

            c.setFillColor(colors.white)
            c.setFont("Helvetica-Bold", 17)
            c.drawString(34*mm, page_h-16*mm, "ELECTRO-DIP")
            c.setFont("Helvetica-Bold", 9)
            c.drawString(34*mm, page_h-23*mm, "MACHINE DAILY PRODUCTION SLIP")
            c.setFont("Helvetica-Bold", 7)
            c.drawRightString(
                page_w-10*mm, page_h-14*mm,
                f"Slip No.: MS-{slip_date:%Y%m%d}-{slip_index:03d}"
            )
            c.drawRightString(
                page_w-10*mm, page_h-21*mm,
                f"Date: {slip_date:%d-%b-%Y}"
            )
            plan_ids = sorted({
                str(row.get("plan_id") or "")
                for row in group_rows
                if row.get("plan_id")
            })
            plan_id_text = ", ".join(plan_ids[:3])
            if len(plan_ids) > 3:
                plan_id_text += f" +{len(plan_ids)-3} more"
            c.drawRightString(
                page_w-10*mm, page_h-27*mm,
                f"Plan ID: {plan_id_text}"
            )

            def box(x, y, w, h, label, value, label_fill):
                label_w = w * 0.36
                c.setStrokeColor(colors.HexColor("#7F8C8D"))
                c.setLineWidth(0.45)
                c.setFillColor(label_fill)
                c.rect(x, y, label_w, h, stroke=1, fill=1)
                c.setFillColor(colors.white)
                c.rect(x+label_w, y, w-label_w, h, stroke=1, fill=1)
                c.setFillColor(colors.HexColor("#1F1F1F"))
                c.setFont("Helvetica-Bold", 6.5)
                c.drawString(x+1.5*mm, y+h/2-2, label)
                c.setFont("Helvetica-Bold", 8)
                c.drawString(x+label_w+1.5*mm, y+h/2-2, str(value or ""))

            info_y = page_h - 42*mm
            shift_values = sorted({
                str(row.get("shift_name") or "")
                for row in group_rows
                if row.get("shift_name")
            })
            shift_text = ", ".join(shift_values)
            box(8*mm, info_y, 90*mm, 9*mm, "Machine", machine, colors.HexColor("#D9EAF7"))
            box(101*mm, info_y, 55*mm, 9*mm, "Shift", shift_text, colors.HexColor("#E4DFEC"))
            box(159*mm, info_y, 62*mm, 9*mm, "Operator", "", colors.HexColor("#E2F0D9"))
            box(224*mm, info_y, 65*mm, 9*mm, "Supervisor", "", colors.HexColor("#FFF2CC"))

            headers = [
                "Barcode / Plan ID", "Sr", "Start", "End", "Customer",
                "Part No.", "Operation", "Plan Qty", "Actual Qty",
                "Rejected Qty", "Priority", "Status", "Remarks"
            ]
            # Fixed widths are tuned for A4 landscape and match the approved
            # slip layout. The first column contains both barcode and Plan ID.
            widths = [48, 7, 13, 13, 25, 20, 34, 15, 17, 18, 14, 15, 22]
            widths = [w*mm for w in widths]
            row_h = 10.5*mm
            table_x = 8*mm
            table_y = info_y - 6*mm - row_h

            x = table_x
            for header, width in zip(headers, widths):
                c.setFillColor(colors.HexColor("#1F4E78"))
                c.setStrokeColor(colors.HexColor("#7F8C8D"))
                c.rect(x, table_y, width, row_h, stroke=1, fill=1)
                c.setFillColor(colors.white)
                c.setFont("Helvetica-Bold", 6.2)
                c.drawCentredString(x+width/2, table_y+2.5*mm, header)
                x += width

            max_rows = 12
            shown = group_rows[:max_rows]
            total_planned = 0.0

            for idx, row in enumerate(shown, start=1):
                y = table_y - idx*row_h
                start_text = ""
                end_text = ""
                if row.get("production_start_datetime"):
                    start_text = datetime.fromisoformat(
                        row["production_start_datetime"]
                    ).strftime("%H:%M")
                if row.get("production_end_datetime"):
                    end_text = datetime.fromisoformat(
                        row["production_end_datetime"]
                    ).strftime("%H:%M")

                values = [
                    row.get("plan_id", ""), idx, start_text, end_text,
                    row.get("customer_name", ""), row.get("part_name", ""),
                    row.get("operation_name", ""),
                    float(row.get("planned_qty") or 0), "", "",
                    row.get("priority", ""), "", ""
                ]
                total_planned += float(row.get("planned_qty") or 0)

                x = table_x
                for col_idx, (value, width) in enumerate(zip(values, widths)):
                    fill = colors.white if idx % 2 else colors.HexColor("#F2F5F7")
                    if col_idx == 8:
                        fill = colors.HexColor("#E2F0D9")
                    elif col_idx == 9:
                        fill = colors.HexColor("#FCE4D6")
                    c.setFillColor(fill)
                    c.setStrokeColor(colors.HexColor("#7F8C8D"))
                    c.rect(x, y, width, row_h, stroke=1, fill=1)

                    if col_idx == 0:
                        # Real Code 128 barcode immediately before Plan ID.
                        plan_id = str(value or "")
                        barcode = code128.Code128(
                            plan_id,
                            barHeight=5.6*mm,
                            barWidth=0.22,
                            humanReadable=False,
                        )
                        barcode_x = x + 1.0*mm
                        barcode_y = y + 2.2*mm
                        c.setFillColor(colors.black)
                        barcode.drawOn(c, barcode_x, barcode_y)

                        # Visible Plan ID remains beside the barcode as a
                        # manual fallback when a scanner is unavailable.
                        c.setFillColor(colors.HexColor("#1F1F1F"))
                        c.setFont("Helvetica-Bold", 4.4)
                        c.drawString(x + 31.5*mm, y + 4.4*mm, plan_id)
                    else:
                        c.setFillColor(colors.HexColor("#1F1F1F"))
                        c.setFont(
                            "Helvetica-Bold"
                            if col_idx in (4, 5, 6, 7)
                            else "Helvetica",
                            5.8,
                        )
                        text = str(value or "")
                        max_chars = 22
                        if len(text) > max_chars:
                            text = text[:max_chars-2] + ".."
                        if col_idx in (4, 6, 12):
                            c.drawString(x+1.2*mm, y+3.4*mm, text)
                        else:
                            c.drawCentredString(x+width/2, y+3.4*mm, text)
                    x += width

            totals_y = table_y - (max(len(shown), 1)+1)*row_h - 4*mm
            box(8*mm, totals_y, 66*mm, 9*mm, "Total Planned Qty", f"{total_planned:g}", colors.HexColor("#D9EAF7"))
            box(77*mm, totals_y, 66*mm, 9*mm, "Total Actual Qty", "", colors.HexColor("#E2F0D9"))
            box(146*mm, totals_y, 66*mm, 9*mm, "Total Rejected Qty", "", colors.HexColor("#F4CCCC"))
            box(215*mm, totals_y, 74*mm, 9*mm, "Machine Utilization", "____ %", colors.HexColor("#FFF2CC"))

            remarks_y = totals_y - 18*mm
            c.setFillColor(colors.white)
            c.setStrokeColor(colors.HexColor("#7F8C8D"))
            c.rect(8*mm, remarks_y, 195*mm, 15*mm, stroke=1, fill=1)
            c.setFillColor(colors.HexColor("#1F4E78"))
            c.setFont("Helvetica-Bold", 6.4)
            c.drawString(10*mm, remarks_y+11*mm, "SUPERVISOR / QUALITY REMARKS")
            c.setStrokeColor(colors.HexColor("#BDC3C7"))
            c.line(10*mm, remarks_y+7*mm, 200*mm, remarks_y+7*mm)
            c.line(10*mm, remarks_y+3.5*mm, 200*mm, remarks_y+3.5*mm)

            c.setFillColor(colors.white)
            c.setStrokeColor(colors.HexColor("#7F8C8D"))
            c.rect(206*mm, remarks_y, 83*mm, 15*mm, stroke=1, fill=1)
            c.setFillColor(colors.HexColor("#2C3E50"))
            c.setFont("Helvetica", 6.2)
            c.drawString(209*mm, remarks_y+10.5*mm, "Status:")
            c.drawString(226*mm, remarks_y+10.5*mm, "[ ] Completed")
            c.drawString(261*mm, remarks_y+10.5*mm, "[ ] Hold")
            c.drawString(226*mm, remarks_y+4.5*mm, "[ ] Running")
            c.drawString(261*mm, remarks_y+4.5*mm, "[ ] Rework")

            c.setFont("Helvetica", 6.4)
            c.drawString(10*mm, 9*mm, "Operator Sign: ______________________________")
            c.drawCentredString(page_w/2, 9*mm, "Quality Sign: ______________________________")
            c.drawRightString(page_w-10*mm, 9*mm, "Supervisor Sign: ______________________________")

            if len(group_rows) > max_rows:
                c.setFillColor(colors.HexColor("#C0392B"))
                c.setFont("Helvetica-Bold", 6)
                c.drawRightString(
                    page_w-10*mm, remarks_y-4*mm,
                    f"Additional rows not shown: {len(group_rows)-max_rows}"
                )
            c.showPage()

        c.save()
    def export_wip_excel(self, output_path):
        wb = xlsxwriter.Workbook(output_path)
        title = wb.add_format({"bold": True, "font_color": "white", "bg_color": "#1F4E78", "font_size": 14, "align": "center"})
        header = wb.add_format({"bold": True, "font_color": "white", "bg_color": "#5B9BD5", "border": 1, "align": "center", "text_wrap": True})
        cell = wb.add_format({"border": 1})
        qty = wb.add_format({"border": 1, "num_format": "0.00"})
        note = wb.add_format({"border": 1, "text_wrap": True})

        def write_sheet(name, records, sheet_title):
            ws = wb.add_worksheet(name[:31])
            if not records:
                ws.write(0, 0, sheet_title, title)
                ws.write(2, 0, "No data")
                return
            columns=list(records[0].keys())
            ws.merge_range(0,0,0,len(columns)-1,sheet_title,title)
            for c,col in enumerate(columns):
                ws.write(1,c,col,header)
            for r,record in enumerate(records,start=2):
                for c,col in enumerate(columns):
                    value=record.get(col)
                    fmt=qty if isinstance(value,float) else note if col in ("calculation_rule","note","remarks") else cell
                    ws.write(r,c,value,fmt)
            ws.autofilter(1,0,len(records)+1,len(columns)-1)
            ws.freeze_panes(2,0)
            for c,col in enumerate(columns):
                width=34 if col in ("calculation_rule","operation_name","remarks","note") else max(14,min(38,len(col)+5))
                ws.set_column(c,c,width)

        write_sheet("Schedule_Line_Calculation", self.schedule_line_calculation_rows(), "SCHEDULE-LINE PLAN QUANTITY CALCULATION")
        write_sheet("Process_WIP_Allocation", self.process_schedule_wip_rows(), "PROCESS-BOM-WISE WIP ALLOCATION")
        write_sheet("WIP_Summary", self.wip_rows(), "CURRENT PHYSICAL WIP SUMMARY")
        write_sheet("Process_WIP_Report", self.process_wip_report_rows(), "PROCESS-WISE WIP REPORT WITH AGEING")
        write_sheet("Machine_WIP", self.machine_wip_rows(), "MACHINE-WISE WIP")
        write_sheet("Part_WIP", self.part_wip_rows(), "PART-WISE WIP")
        write_sheet("Customer_WIP", self.customer_wip_rows(), "CUSTOMER-WISE WIP")
        write_sheet("Dispatch_History", self.dispatch_history_rows(), "DISPATCH HISTORY")
        write_sheet("Production_Plan", [dict(r) for r in self.plan_rows()], "TRANSPORTATION-BATCH PRODUCTION PLAN")
        write_sheet("Operator_Entries", [dict(r) for r in self.operator_entries()], "PERSISTENT OPERATOR ENTRIES")
        logic=wb.add_worksheet("Logic")
        logic.set_column("A:A",30); logic.set_column("B:B",95)
        logic.write("A1","Rule",header); logic.write("B1","Application",header)
        rules=[
            ("First schedule line","Plan Qty = Demand + Minimum Stock Level - Opening Available Stock/WIP."),
            ("Further schedule lines","Plan Qty = Demand + Minimum Stock Level - Previous Schedule Carry-Forward WIP."),
            ("Process WIP","Each Process BOM sequence has its own WIP pool. WIP is consumed FIFO and cannot be reused."),
            ("Dispatch","Highest Process BOM sequence is Dispatch. Only dispatch good quantity closes customer demand."),
            ("Transportation batch","Net process plan is split into transportation batches. The final lot may be smaller."),
            ("Persistence","Operator entries remain saved until Clear Previous Entries is used."),
        ]
        for r,row in enumerate(rules,start=1): logic.write_row(r,0,row,cell)
        wb.close()

